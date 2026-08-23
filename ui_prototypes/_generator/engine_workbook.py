"""Build-time full audited Excel workbook for DNL, sourced from the production engine.

Standalone-UI feature #2 ("download EVERYTHING to Excel"). A static shareable HTML
cannot call the Python engine at runtime, so the workbook is pre-generated here at
build time and base64-embedded in the page; the download button serves the bytes.

Discipline (standing rule 1): every yellow cell on the Assumptions sheet is a DATA
input; every other cell on every other sheet is an Excel FORMULA that links back to
Assumptions (or to a prior derived cell), so the whole model can be traced, audited
and flexed by hand in Excel. Values are NOT Python-computed and pasted — the engine
runs only to pull the yellow-cell inputs and to CHECK the formulas tie (see the
verify script). The industry-archetype baseline and the company-position offset are
kept as separate input rows, with the company/ scenario value derived by formula.

Nothing here hardcodes a domain number: all inputs are read live from the loaded
data via the translator, so this stays downstream of the one source of truth.
"""

from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_ROOT = Path(__file__).resolve().parents[2]
import sys as _sys
_sys.path.insert(0, str(_ROOT / "src"))
from vcc_valuations.translator import (  # noqa: E402
    load_inputs, build_engine_inputs_from_data, build_wacc_from_inputs,
    tax_bridge_from_data, wacc_build_from_data, equity_bridge_from_data,
    revenue_growth_chain_from_data, _geographic_regions,
)
from vcc_valuations.dcf.fcf_engine import FcfEngine  # noqa: E402

# scenario order (display) — central case is muddle_through
SCEN = [
    ("orderly_convergence", "Orderly Convergence"),
    ("muddle_through", "Muddle Through"),
    ("ai_productivity_lag", "AI Productivity Lag"),
    ("fragmentation", "Fragmentation"),
    ("disorderly_climate_crystallisation", "Disorderly Climate"),
    ("stagflation_persists", "Stagflation Persists"),
]
CENTRAL = "muddle_through"
PERIODS = ["Stub", "Y1", "Y2", "Y3", "Y4", "Y5"]  # p=0 stub

# ---- styling ----
YELLOW = PatternFill("solid", fgColor="FFF2CC")
BLUEFONT = Font(color="1F4E78")
HDR = Font(bold=True, size=12)
SUB = Font(bold=True, color="404040")
BOLD = Font(bold=True)
NOTE = Font(italic=True, color="808080", size=9)
PCT2 = "0.00%"
PCT3 = "0.000%"
NUM1 = "#,##0.0"
NUM0 = "#,##0"
MONEY = "#,##0.0"
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(bottom=thin)


def _load_central():
    return load_inputs(_ROOT, CENTRAL, "industrial_explosives", "dnl")


def _col(i):  # 1-based -> letter
    return get_column_letter(i)


class Book:
    def __init__(self):
        self.wb = Workbook()
        self.wb.calculation.fullCalcOnLoad = True
        self.ref = {}  # key -> "'Sheet'!$C$R"

    # ---- Assumptions -------------------------------------------------------
    def build_assumptions(self, inp):
        ws = self.wb.active
        ws.title = "Assumptions"
        ws.sheet_properties.tabColor = "FFD966"
        craw = inp["company_raw"]; nb = craw["normalised_baseline"]; fin = inp["financials"]
        rgc = nb["revenue_growth_chain"]; shared = rgc["shared"]
        istr = shared["industry_structure"]; coff = shared["company_offset"]
        ov = nb["engine_overlays"]; base = ov["baseline"]
        tb = nb["tax_bridge"]; rr = nb["equity_bridge_run_rates"]
        w = build_wacc_from_inputs(inp)
        geo = _geographic_regions(craw)

        r = [1]
        def line(label, value=None, key=None, fmt=None, style=None, yellow=True):
            ws.cell(r[0], 1, label)
            if style:
                ws.cell(r[0], 1).font = style
            if value is not None:
                c = ws.cell(r[0], 2, value)
                if yellow:
                    c.fill = YELLOW; c.font = BLUEFONT
                if fmt:
                    c.number_format = fmt
                if key:
                    self.ref[key] = f"'Assumptions'!$B${r[0]}"
            r[0] += 1
        def blank():
            r[0] += 1

        ws.cell(r[0], 1, "DNL — Assumptions (yellow cells are the only inputs)"); ws.cell(r[0], 1).font = HDR; r[0] += 1
        ws.cell(r[0], 1, "Every other sheet links here by formula (standing rule 1). Sourced from the production engine's data files."); ws.cell(r[0], 1).font = NOTE; r[0] += 1
        blank()

        # --- Macro drivers by scenario (table: scenarios as columns C..H) ---
        ws.cell(r[0], 1, "Macro drivers by scenario"); ws.cell(r[0], 1).font = SUB; r[0] += 1
        hdr_row = r[0]
        ws.cell(hdr_row, 1, "Driver"); ws.cell(hdr_row, 1).font = BOLD
        for j, (sid, nm) in enumerate(SCEN):
            c = ws.cell(hdr_row, 3 + j, nm); c.font = BOLD; c.alignment = Alignment(horizontal="right", wrap_text=True)
        r[0] += 1
        macro_rows = {}
        drivers = [
            ("dm_inflation", "DM inflation", PCT2, lambda s: rgc["by_scenario"][s]["macro"]["dm_inflation"]),
            ("global_mining_real_growth", "Global mining real growth", PCT2, lambda s: rgc["by_scenario"][s]["macro"]["global_mining_real_growth"]),
            ("gas_price_growth", "Gas price growth", PCT2, lambda s: rgc["by_scenario"][s]["macro"]["gas_price_growth"]),
            ("terminal_growth", "Terminal growth g", PCT2, lambda s: ov["by_scenario"][s]["terminal_growth"]),
            ("margin_delta_pp", "Margin delta (parallel shift, pp)", PCT2, lambda s: ov["by_scenario"][s].get("margin_delta_pp", 0.0)),
            ("capex_delta_pp", "Capex delta (parallel shift, pp)", PCT2, lambda s: ov["by_scenario"][s].get("capex_delta_pp", 0.0)),
        ]
        for key, label, fmt, fn in drivers:
            ws.cell(r[0], 1, label)
            macro_rows[key] = r[0]
            for j, (sid, nm) in enumerate(SCEN):
                c = ws.cell(r[0], 3 + j, fn(sid)); c.fill = YELLOW; c.font = BLUEFONT; c.number_format = fmt
                self.ref[f"macro:{key}:{sid}"] = f"'Assumptions'!${_col(3 + j)}${r[0]}"
            r[0] += 1
        blank()

        # --- Industry structure (shared, scenario-invariant) ---
        ws.cell(r[0], 1, "Industry-structure coefficients (archetype baseline, shared)"); ws.cell(r[0], 1).font = SUB; r[0] += 1
        line("Volume coefficient a (mining beta)", istr["volume_coefficient_a"], "a", NUM1 + "0")
        line("Volume constant b", istr["volume_constant_b"], "b", PCT3)
        line("Pricing weight — inflation", istr["pricing_weight_inflation"], "w_infl", NUM1 + "0")
        line("Pricing weight — gas", istr["pricing_weight_gas"], "w_gas", NUM1 + "0")
        line("Productivity sharing", istr["productivity_sharing"], "prod", PCT2)
        blank()

        # --- Company offset (shared) ---
        ws.cell(r[0], 1, "Company-position offset (shared)"); ws.cell(r[0], 1).font = SUB; r[0] += 1
        line("EM growth premium", coff["em_growth_premium"], "em_prem", NUM1 + "0")
        ff = coff["five_forces_offset"]
        line("Five Forces — rivalry (competitive position)", ff["rivalry_competitive_position"], "ff_riv", PCT2)
        line("Five Forces — rivalry (product mix)", ff["rivalry_product_mix"], "ff_mix", PCT2)
        line("Five Forces — new entrants (pipeline uplift)", ff["new_entrants_pipeline_uplift"], "ff_ent", PCT2)
        line("Five Forces — buyer/supplier/substitutes", ff["buyer_supplier_substitutes"], "ff_bss", PCT2)
        blank()

        # --- Geographic concentration & tax jurisdictions ---
        ws.cell(r[0], 1, "Geographic concentration & statutory tax"); ws.cell(r[0], 1).font = SUB; r[0] += 1
        gh = r[0]; ws.cell(gh, 1, "Region"); ws.cell(gh, 2, "Revenue weight"); ws.cell(gh, 3, "Statutory rate"); ws.cell(gh, 4, "Developed-market?")
        for cc in (1, 2, 3, 4):
            ws.cell(gh, cc).font = BOLD
        r[0] += 1
        dm_regions = set(coff["developed_market_regions"])
        geo_rows = {}
        for g in geo:
            name = g["geo"]; ws.cell(r[0], 1, name)
            cw = ws.cell(r[0], 2, g["share_of_revenue"]); cw.fill = YELLOW; cw.font = BLUEFONT; cw.number_format = PCT2
            self.ref[f"geo_w:{name}"] = f"'Assumptions'!$B${r[0]}"
            cr = ws.cell(r[0], 3, tb["statutory_rate_by_region"][name]); cr.fill = YELLOW; cr.font = BLUEFONT; cr.number_format = PCT2
            self.ref[f"geo_stat:{name}"] = f"'Assumptions'!$C${r[0]}"
            dm = "Yes" if name in dm_regions else "No"
            cd = ws.cell(r[0], 4, dm); cd.fill = YELLOW; cd.font = BLUEFONT
            self.ref[f"geo_dm:{name}"] = f"'Assumptions'!$D${r[0]}"
            geo_rows[name] = r[0]
            r[0] += 1
        self.geo_names = [g["geo"] for g in geo]
        line("Effective tax rate (stub)", tb["effective_tax_rate"], "eff_tax", PCT2)
        # glide fractions
        ws.cell(r[0], 1, "Tax glide fractions (Y1..Y5)"); 
        for i, fr in enumerate(tb["glide_fractions"]):
            c = ws.cell(r[0], 3 + i, fr); c.fill = YELLOW; c.font = BLUEFONT; c.number_format = NUM1 + "00"
            self.ref[f"glide_frac:{i}"] = f"'Assumptions'!${_col(3 + i)}${r[0]}"
        r[0] += 1
        blank()

        # --- WACC inputs ---
        ws.cell(r[0], 1, "Discount rate (WACC) inputs"); ws.cell(r[0], 1).font = SUB; r[0] += 1
        line("Risk-free rate Rf", w.risk_free_rate, "rf", PCT2)
        line("Equity risk premium ERP", w.equity_risk_premium, "erp", PCT2)
        line("Beta (triangulated, ratified)", w.beta, "beta", NUM1 + "0")
        line("Pre-tax cost of debt Rd", w.cost_of_debt_pretax, "rd", PCT2)
        line("Tax rate (for after-tax Kd)", w.tax_rate, "tax_wacc", PCT2)
        line("Equity market value E", w.equity_market_value, "E", NUM0)
        line("Debt market value D", w.debt_market_value, "D", NUM0)
        blank()

        # --- Operating build baseline (per-year vectors, shared) ---
        ws.cell(r[0], 1, "Operating build baseline (shared; scenario applies parallel shift)"); ws.cell(r[0], 1).font = SUB; r[0] += 1
        line("Base-year revenue (continuing ops)", nb["base_year_revenue"], "base_rev", NUM0)
        line("Base EBIT margin", base["base_ebit_margin"], "base_margin", PCT2)
        line("D&A % of revenue", base["da_pct_revenue"], "da_pct", PCT2)
        line("Capex % — stub", base["capex_pct_stub"], "capex_stub", PCT2)
        line("Stub years", nb["stub_years"], "stub_years", NUM1 + "00")
        line("Horizon years", nb["horizon_years"], "horizon", "0")
        # margin transformation vector
        ws.cell(r[0], 1, "Margin transformation (Y1..Y5, pp)")
        for i, v in enumerate(base["margin_transformation"]):
            c = ws.cell(r[0], 3 + i, v); c.fill = YELLOW; c.font = BLUEFONT; c.number_format = PCT2
            self.ref[f"mt:{i}"] = f"'Assumptions'!${_col(3 + i)}${r[0]}"
        r[0] += 1
        ws.cell(r[0], 1, "Margin gas roll-off (Y1..Y5, pp)")
        for i, v in enumerate(base["margin_gas_rolloff"]):
            c = ws.cell(r[0], 3 + i, v); c.fill = YELLOW; c.font = BLUEFONT; c.number_format = PCT2
            self.ref[f"gas:{i}"] = f"'Assumptions'!${_col(3 + i)}${r[0]}"
        r[0] += 1
        ws.cell(r[0], 1, "Capex % (Y1..Y5)")
        for i, v in enumerate(base["capex_pct"]):
            c = ws.cell(r[0], 3 + i, v); c.fill = YELLOW; c.font = BLUEFONT; c.number_format = PCT2
            self.ref[f"capex:{i}"] = f"'Assumptions'!${_col(3 + i)}${r[0]}"
        r[0] += 1
        blank()

        # --- Working capital (working_capital_treatment.md) ---
        # The four balance-sheet lines and revenue are the ONLY inputs; the
        # year intensity, the clean-year average, the rounding and the applied
        # figure are all Excel formulas, so the derivation Stephen ratified can
        # be re-run and flexed by hand rather than taken on trust.
        wci = nb["working_capital_intensity"]
        hist = {row["fiscal_year"]: row for row in fin["working_capital_history"]}
        ws.cell(r[0], 1, "Working capital — derived intensity (never an input)"); ws.cell(r[0], 1).font = SUB; r[0] += 1
        wc_hdr = r[0]
        for cc, lbl in enumerate(
            ["Clean year", "Total current assets", "Less: cash & equivalents",
             "Total current liabilities", "Less: interest-bearing (incl. current leases)",
             "Revenue", "NCWC / revenue"], start=1
        ):
            c = ws.cell(wc_hdr, cc, lbl); c.font = BOLD
            c.alignment = Alignment(horizontal="right", wrap_text=True)
        r[0] += 1
        level_cells = []
        for fy in wci["clean_years"]:
            row = hist[fy]
            ws.cell(r[0], 1, fy)
            for cc, key in enumerate(
                ["total_current_assets", "cash_and_short_term_investments",
                 "total_current_liabilities", "interest_bearing_current_debt", "revenue"],
                start=2,
            ):
                c = ws.cell(r[0], cc, row[key]); c.fill = YELLOW; c.font = BLUEFONT; c.number_format = NUM1
            ca, cash, cl, ibd, rev = (f"${_col(cc)}${r[0]}" for cc in range(2, 7))
            c = ws.cell(r[0], 7, f"=(({ca}-{cash})-({cl}-{ibd}))/{rev}"); c.number_format = PCT2
            level_cells.append(f"$G${r[0]}")
            r[0] += 1
        line("Average of clean years", None, None, PCT2)
        ws.cell(r[0] - 1, 2, "=AVERAGE(" + ",".join(level_cells) + ")").number_format = PCT2
        self.ref["wc_avg"] = f"'Assumptions'!$B${r[0] - 1}"
        avg = f"$B${r[0] - 1}"
        line("Rounded (protocol: nearest 5pp)", None, None, PCT2)
        step = 1.0 / 20.0  # ssot-allow: the D-29 rounding bucket, not a register value
        ws.cell(r[0] - 1, 2, f"=ROUND({avg}/{step},0)*{step}").number_format = PCT2
        rounded = f"$B${r[0] - 1}"
        applied_src = rounded
        if "rounding_override" in wci:
            line("Override — raw figure held (single observation)",
                 wci["rounding_override"], "wc_override", PCT2)
            applied_src = self.ref["wc_override"].split("!")[1]
        line("Applied working-capital intensity", None, None, PCT2, style=BOLD)
        c = ws.cell(r[0] - 1, 2, f"={applied_src}")
        c.number_format = PCT2; c.font = BOLD
        self.ref["wc_int"] = f"'Assumptions'!$B${r[0] - 1}"
        tr = nb["terminal_reinvestment"]
        self.terminal_mode = tr["mode"]
        rule = tr["capex_rule"]
        src = "da_pct" if rule == "equals_da" else f"capex:{len(base['capex_pct']) - 1}"
        line(f"Terminal capex % of revenue (rule: {rule})", None, None, PCT2)
        ws.cell(r[0] - 1, 2, "=" + self.ref[src].split("!")[1]).number_format = PCT2
        self.ref["t_capex"] = f"'Assumptions'!$B${r[0] - 1}"
        blank()

        # --- Equity bridge run-rates ---
        ws.cell(r[0], 1, "Equity bridge run-rates & anchors"); ws.cell(r[0], 1).font = SUB; r[0] += 1
        line("Net debt at anchor (31 Mar 2026, §5.3)", fin["derived_metrics"]["net_debt"], "nd_anchor", NUM1)
        line("Period A days", rr["period_a_days"], "pa_days", "0")
        line("Operating cash flow run-rate (annual)", rr["operating_cash_flow_run_rate"], "ocf_rr", NUM0)
        line("Capex run-rate (annual)", rr["capex_run_rate"], "capex_rr", NUM0)
        line("AASB 16 lease liabilities", rr["lease_liabilities"], "leases", NUM1)
        line("Shares outstanding (m)", fin["share_statistics"]["shares_outstanding"] / 1_000_000, "shares", NUM0)
        line("Market reference price", rr["market_reference_price"], "mkt", NUM1 + "0")
        # equity-bridge adjustments net (derived elsewhere; here as one audited input line with detail on its own sheet)
        blank()

        ws.column_dimensions["A"].width = 44
        for cc in "BCDEFGH":
            ws.column_dimensions[cc].width = 13
        self.macro_rows = macro_rows

    # ---- Revenue growth chain (six scenarios as columns) -------------------
    def build_revenue(self, inp):
        ws = self.wb.create_sheet("Revenue growth")
        R = self.ref
        ws.cell(1, 1, "Revenue-growth chain (methodology §11) — company nominal growth by scenario"); ws.cell(1, 1).font = HDR
        ws.cell(2, 1, "Industry baseline + company offset, both by formula off Assumptions."); ws.cell(2, 1).font = NOTE
        hr = 4
        ws.cell(hr, 1, "Derived row (workbook cell)"); ws.cell(hr, 1).font = BOLD
        for j, (sid, nm) in enumerate(SCEN):
            c = ws.cell(hr, 3 + j, nm); c.font = BOLD; c.alignment = Alignment(horizontal="right", wrap_text=True)
        row = {}
        def put(cell_key, label, fmt, formula_fn):
            rr = hr + 1 + len(row)
            ws.cell(rr, 1, label)
            for j, (sid, nm) in enumerate(SCEN):
                col = _col(3 + j)
                c = ws.cell(rr, 3 + j, formula_fn(sid, col)); c.number_format = fmt
            row[cell_key] = rr
        # geo weights (DM/EM) come from Assumptions (same for all scenarios)
        dm_cells = "+".join(R[f"geo_w:{g}"] for g in self.geo_names if R[f"geo_dm:{g}"])  # placeholder, replaced below
        dm_names = [g for g in self.geo_names]
        # compute DM/EM sums using the Yes/No flags -> use SUMIF-like explicit sum
        dm_ref = "+".join(R[f"geo_w:{g}"] for g in dm_names)  # not used
        # We know DM = developed regions; build explicit sums
        # (developed flagged by geo_dm == 'Yes'); we captured membership when writing Assumptions
        put("B25", "B25 Industry volume growth",
            PCT3, lambda s, c: f"={R['a']}*{R[f'macro:global_mining_real_growth:{s}']}+{R['b']}")
        put("B29", "B29 Industry pricing growth",
            PCT3, lambda s, c: f"={R['w_infl']}*{R[f'macro:dm_inflation:{s}']}+{R['w_gas']}*{R[f'macro:gas_price_growth:{s}']}+{R['prod']}")
        put("B30", "B30 Industry nominal growth",
            PCT3, lambda s, c: f"=(1+{c}{row['B25']})*(1+{c}{row['B29']})-1")
        # DM/EM weighting: sum revenue weights of developed / non-developed regions
        dm_sum = "+".join(R[f"geo_w:{g}"] for g in self._dm_regions)
        em_sum = "+".join(R[f"geo_w:{g}"] for g in self._em_regions) or "0"
        put("B33", "B33 DM weighting", PCT2, lambda s, c: f"={dm_sum}")
        put("B34", "B34 EM weighting", PCT2, lambda s, c: f"={em_sum}")
        put("B36", "B36 Geographic-mix multiplier",
            NUM1 + "00", lambda s, c: f"={c}{row['B33']}+{c}{row['B34']}*{R['em_prem']}")
        put("B41", "B41 Net company-position offset",
            PCT3, lambda s, c: f"={R['ff_riv']}+{R['ff_mix']}+{R['ff_ent']}+{R['ff_bss']}")
        put("B42", "B42 Company nominal revenue growth",
            PCT3, lambda s, c: f"={c}{row['B30']}*{c}{row['B36']}+{c}{row['B41']}")
        self.rev_rows = row
        # record per-scenario B42 ref
        for j, (sid, nm) in enumerate(SCEN):
            self.ref[f"rev_growth:{sid}"] = f"'Revenue growth'!{_col(3 + j)}{row['B42']}"
        ws.column_dimensions["A"].width = 40
        for cc in "CDEFGH":
            ws.column_dimensions[cc].width = 14

    # ---- WACC build --------------------------------------------------------
    def build_wacc(self):
        ws = self.wb.create_sheet("WACC build")
        R = self.ref
        ws.cell(1, 1, "WACC build (single discount rate, all scenarios)"); ws.cell(1, 1).font = HDR
        rows = [
            ("B8", "Cost of equity Re = Rf + beta*ERP", PCT3, f"={R['rf']}+{R['beta']}*{R['erp']}"),
            ("B13", "After-tax cost of debt = Rd*(1-tax)", PCT3, f"={R['rd']}*(1-{R['tax_wacc']})"),
            ("B18", "Enterprise value V = E + D", NUM1, f"={R['E']}+{R['D']}"),
        ]
        rr = 3
        cellrow = {}
        for k, label, fmt, formula in rows:
            ws.cell(rr, 1, label); c = ws.cell(rr, 2, formula); c.number_format = fmt
            cellrow[k] = rr; rr += 1
        ws.cell(rr, 1, "E/V"); c = ws.cell(rr, 2, f"={R['E']}/B{cellrow['B18']}"); c.number_format = PCT2; cellrow["B19"] = rr; rr += 1
        ws.cell(rr, 1, "D/V"); c = ws.cell(rr, 2, f"={R['D']}/B{cellrow['B18']}"); c.number_format = PCT2; cellrow["B20"] = rr; rr += 1
        ws.cell(rr, 1, "WACC = (E/V)*Re + (D/V)*Rd(after-tax)"); ws.cell(rr, 1).font = BOLD
        c = ws.cell(rr, 2, f"=B{cellrow['B19']}*B{cellrow['B8']}+B{cellrow['B20']}*B{cellrow['B13']}")
        c.number_format = PCT3; c.font = BOLD; cellrow["B23"] = rr
        self.ref["wacc"] = f"'WACC build'!$B${cellrow['B23']}"
        ws.column_dimensions["A"].width = 42; ws.column_dimensions["B"].width = 14

    # ---- Tax bridge --------------------------------------------------------
    def build_tax(self):
        ws = self.wb.create_sheet("Tax bridge")
        R = self.ref
        ws.cell(1, 1, "Tax bridge — effective rate gliding to blended statutory"); ws.cell(1, 1).font = HDR
        rr = 3
        contrib_rows = []
        for g in self.geo_names:
            ws.cell(rr, 1, f"{g} contribution = weight x statutory")
            c = ws.cell(rr, 2, f"={R[f'geo_w:{g}']}*{R[f'geo_stat:{g}']}"); c.number_format = PCT3
            contrib_rows.append(rr); rr += 1
        ws.cell(rr, 1, "Blended statutory rate (D8)"); ws.cell(rr, 1).font = BOLD
        c = ws.cell(rr, 2, "=" + "+".join(f"B{x}" for x in contrib_rows)); c.number_format = PCT3; c.font = BOLD
        d8 = rr; self.ref["blended_stat"] = f"'Tax bridge'!$B${d8}"; rr += 1
        rr += 1
        ws.cell(rr, 1, "Applied tax glide (effective + (blended-effective)*fraction)"); ws.cell(rr, 1).font = SUB; rr += 1
        self.tax_glide_rows = []
        for i in range(5):
            ws.cell(rr, 1, f"FY{27+i} applied tax rate (Y{i+1})")
            c = ws.cell(rr, 2, f"={R['eff_tax']}+(B{d8}-{R['eff_tax']})*{R[f'glide_frac:{i}']}"); c.number_format = PCT3
            self.ref[f"tax_glide:{i}"] = f"'Tax bridge'!$B${rr}"
            self.tax_glide_rows.append(rr); rr += 1
        ws.column_dimensions["A"].width = 52; ws.column_dimensions["B"].width = 14

    # ---- DCF build (six scenarios as columns) ------------------------------
    def build_dcf(self):
        ws = self.wb.create_sheet("DCF build")
        R = self.ref
        ws.cell(1, 1, "DCF build — FCFF per year, all six scenarios (formula-linked)"); ws.cell(1, 1).font = HDR
        ws.cell(2, 1, "Single WACC across scenarios; margin/capex deltas applied as a parallel shift; stub is a partial year."); ws.cell(2, 1).font = NOTE
        hr = 4
        ws.cell(hr, 1, "Line"); ws.cell(hr, 1).font = BOLD
        for j, (sid, nm) in enumerate(SCEN):
            c = ws.cell(hr, 3 + j, nm); c.font = BOLD; c.alignment = Alignment(horizontal="right", wrap_text=True)
        SC = [_col(3 + j) for j in range(len(SCEN))]
        SIDS = [s for s, _ in SCEN]
        rmap = {}
        def block(name):
            ws.cell(hr + 1 + len(rmap.get("_order", [])) if False else 0, 1)  # noop
        rr = [hr + 1]
        def section(label):
            ws.cell(rr[0], 1, label); ws.cell(rr[0], 1).font = SUB; rr[0] += 1
        def prow(key, label, fmt, fn, bold=False):
            ws.cell(rr[0], 1, label)
            if bold:
                ws.cell(rr[0], 1).font = BOLD
            for j, sid in enumerate(SIDS):
                col = SC[j]
                c = ws.cell(rr[0], 3 + j, fn(sid, col, j)); c.number_format = fmt
                if bold:
                    c.font = BOLD
            rmap[key] = rr[0]; rr[0] += 1

        # revenue growth link
        prow("g_rev", "Revenue growth (from Revenue growth sheet)", PCT3,
             lambda s, c, j: f"={R[f'rev_growth:{s}']}")
        # Revenue rows p0..p5
        section("Revenue")
        for p in range(6):
            def f_rev(s, c, j, p=p):
                if p == 0:
                    return f"={R['base_rev']}*{R['stub_years']}"
                return f"={R['base_rev']}*(1+{c}{rmap['g_rev']})^{p}"
            prow(f"rev{p}", f"  {PERIODS[p]} revenue", MONEY, f_rev)
        # EBIT margin
        section("EBIT margin")
        for p in range(6):
            def f_m(s, c, j, p=p):
                if p == 0:
                    return f"={R['base_margin']}"
                i = p - 1
                return f"={R['base_margin']}+({R[f'mt:{i}']}+{R[f'macro:margin_delta_pp:{s}']})+{R[f'gas:{i}']}"
            prow(f"m{p}", f"  {PERIODS[p]} EBIT margin", PCT2, f_m)
        # EBIT
        section("EBIT")
        for p in range(6):
            prow(f"ebit{p}", f"  {PERIODS[p]} EBIT", MONEY,
                 lambda s, c, j, p=p: f"={c}{rmap[f'rev{p}']}*{c}{rmap[f'm{p}']}")
        # applied tax
        section("Applied tax rate")
        for p in range(6):
            def f_t(s, c, j, p=p):
                if p == 0:
                    return f"={R['eff_tax']}"
                return f"={R[f'tax_glide:{p-1}']}"
            prow(f"tax{p}", f"  {PERIODS[p]} applied tax rate", PCT3, f_t)
        # NOPAT
        section("NOPAT = EBIT x (1 - tax)")
        for p in range(6):
            prow(f"nop{p}", f"  {PERIODS[p]} NOPAT", MONEY,
                 lambda s, c, j, p=p: f"={c}{rmap[f'ebit{p}']}*(1-{c}{rmap[f'tax{p}']})")
        # D&A
        section("D&A = revenue x D&A%")
        for p in range(6):
            prow(f"da{p}", f"  {PERIODS[p]} D&A", MONEY,
                 lambda s, c, j, p=p: f"={c}{rmap[f'rev{p}']}*{R['da_pct']}")
        # Capex (negative)
        section("Capex (negative)")
        for p in range(6):
            def f_cx(s, c, j, p=p):
                if p == 0:
                    return f"=-{c}{rmap[f'rev{p}']}*{R['capex_stub']}"
                i = p - 1
                return f"=-{c}{rmap[f'rev{p}']}*({R[f'capex:{i}']}+{R[f'macro:capex_delta_pp:{s}']})"
            prow(f"cx{p}", f"  {PERIODS[p]} capex", MONEY, f_cx)
        # Working capital. NCWC is a stock and scales with the ANNUALISED
        # revenue run-rate, not with the stub's part-year flow — hence the
        # run-rate row, which differs from the revenue row only in the stub.
        section("Working capital (intensity x change in annualised revenue)")
        for p in range(6):
            def f_rrate(s, c, j, p=p):
                exp = f"{R['stub_years']}" if p == 0 else f"{p}"
                return f"={R['base_rev']}*(1+{c}{rmap['g_rev']})^{exp}"
            prow(f"rrate{p}", f"  {PERIODS[p]} revenue run-rate (annualised)", MONEY, f_rrate)
        for p in range(6):
            def f_dwc(s, c, j, p=p):
                prev = R["base_rev"] if p == 0 else f"{c}{rmap[f'rrate{p-1}']}"
                return f"=-{R['wc_int']}*({c}{rmap[f'rrate{p}']}-{prev})"
            prow(f"dwc{p}", f"  {PERIODS[p]} change in working capital", MONEY, f_dwc)
        # FCFF
        section("FCFF = NOPAT + D&A + Capex + ΔWC")
        for p in range(6):
            prow(f"fcff{p}", f"  {PERIODS[p]} FCFF", MONEY,
                 lambda s, c, j, p=p: f"={c}{rmap[f'nop{p}']}+{c}{rmap[f'da{p}']}+{c}{rmap[f'cx{p}']}+{c}{rmap[f'dwc{p}']}")
        # mid-times & discount factors
        section("Discounting (single WACC)")
        for p in range(6):
            def f_mt(s, c, j, p=p):
                if p == 0:
                    return f"={R['stub_years']}/2"
                return f"={R['stub_years']}+{p}-0.5"
            prow(f"midt{p}", f"  {PERIODS[p]} mid-time (yrs)", NUM1 + "00", f_mt)
        for p in range(6):
            prow(f"df{p}", f"  {PERIODS[p]} discount factor", NUM1 + "0000",
                 lambda s, c, j, p=p: f"=1/(1+{R['wacc']})^{c}{rmap[f'midt{p}']}")
        for p in range(6):
            prow(f"pv{p}", f"  {PERIODS[p]} PV of FCFF", MONEY,
                 lambda s, c, j, p=p: f"={c}{rmap[f'fcff{p}']}*{c}{rmap[f'df{p}']}")
        # terminal + EV
        section("Terminal value & enterprise value")
        prow("pv_expl", "PV of explicit FCFF", MONEY,
             lambda s, c, j: "=" + "+".join(f"{c}{rmap[f'pv{p}']}" for p in range(6)))
        if self.terminal_mode == "normalised":
            # Rebuilt from components: Y5 margin and tax, D&A, terminal capex
            # (declared rule), and a working-capital drag of g x intensity.
            # Capitalising Y5 FCFF instead would carry the explicit period's
            # capex and working-capital build into perpetuity.
            prow("tfcff", "Terminal FCFF (normalised reinvestment)", MONEY,
                 lambda s, c, j: (
                     f"={c}{rmap['rev5']}*(1+{R[f'macro:terminal_growth:{s}']})"
                     f"*({c}{rmap['m5']}*(1-{c}{rmap['tax5']})+{R['da_pct']}"
                     f"-{R['t_capex']}-{R[f'macro:terminal_growth:{s}']}*{R['wc_int']})"
                 ))
        else:
            prow("tfcff", "Terminal FCFF = Y5 FCFF x (1+g)", MONEY,
                 lambda s, c, j: f"={c}{rmap['fcff5']}*(1+{R[f'macro:terminal_growth:{s}']})")
        prow("tv", "Terminal value = TFCFF/(WACC-g)", MONEY,
             lambda s, c, j: f"={c}{rmap['tfcff']}/({R['wacc']}-{R[f'macro:terminal_growth:{s}']})")
        prow("tend", "Terminal end time (stub+H)", NUM1 + "00",
             lambda s, c, j: f"={R['stub_years']}+{R['horizon']}")
        prow("tdf", "Terminal discount factor", NUM1 + "0000",
             lambda s, c, j: f"=1/(1+{R['wacc']})^{c}{rmap['tend']}")
        prow("pv_term", "PV of terminal value", MONEY,
             lambda s, c, j: f"={c}{rmap['tv']}*{c}{rmap['tdf']}")
        prow("ev", "Enterprise value (EV)", MONEY,
             lambda s, c, j: f"={c}{rmap['pv_expl']}+{c}{rmap['pv_term']}", bold=True)
        # equity bridge (walk is scenario-invariant; reference Assumptions run-rates)
        section("Equity bridge")
        prow("nd_val", "Less: net debt at valuation date", MONEY,
             lambda s, c, j: f"=-({R['nd_anchor']}-{R['ocf_rr']}*{R['pa_days']}/365+{R['capex_rr']}*{R['pa_days']}/365)")
        prow("adj", "Less: equity-bridge adjustments (net)", MONEY,
             lambda s, c, j: f"=-'Equity bridge'!$B${self.adj_row}")
        prow("leas", "Less: AASB 16 lease liabilities", MONEY,
             lambda s, c, j: f"=-{R['leases']}")
        prow("eq", "Equity value", MONEY,
             lambda s, c, j: f"={c}{rmap['ev']}+{c}{rmap['nd_val']}+{c}{rmap['adj']}+{c}{rmap['leas']}", bold=True)
        prow("vps", "Value per share", NUM1 + "00",
             lambda s, c, j: f"={c}{rmap['eq']}/{R['shares']}", bold=True)
        prow("vsm", "vs market", PCT2,
             lambda s, c, j: f"={c}{rmap['vps']}/{R['mkt']}-1")
        self.dcf_rows = rmap
        for j, sid in enumerate(SCEN):
            pass
        for j, (sid, nm) in enumerate(SCEN):
            self.ref[f"vps:{sid}"] = f"'DCF build'!{SC[j]}{rmap['vps']}"
            self.ref[f"ev:{sid}"] = f"'DCF build'!{SC[j]}{rmap['ev']}"
        ws.column_dimensions["A"].width = 42
        for cc in "CDEFGH":
            ws.column_dimensions[cc].width = 13

    # ---- Equity bridge (central case detail) -------------------------------
    def build_equity(self, inp):
        ws = self.wb.create_sheet("Equity bridge")
        R = self.ref
        ws.cell(1, 1, "Equity bridge — central case (Muddle Through), Period-A net-debt walk (methodology §4/§7)"); ws.cell(1, 1).font = HDR
        rr = 3
        ws.cell(rr, 1, "Net debt at anchor (31 Mar 2026)"); ws.cell(rr, 2, f"={R['nd_anchor']}").number_format = NUM1; b6 = rr; rr += 1
        ws.cell(rr, 1, "less: operating cash flow in Period A"); ws.cell(rr, 2, f"=-{R['ocf_rr']}*{R['pa_days']}/365").number_format = NUM1; b7 = rr; rr += 1
        ws.cell(rr, 1, "plus: capex paid in Period A"); ws.cell(rr, 2, f"={R['capex_rr']}*{R['pa_days']}/365").number_format = NUM1; b8 = rr; rr += 1
        ws.cell(rr, 1, "Net debt at valuation date"); ws.cell(rr, 1).font = BOLD
        c = ws.cell(rr, 2, f"=B{b6}+B{b7}+B{b8}"); c.number_format = NUM1; c.font = BOLD; b11 = rr; rr += 1
        rr += 1
        # adjustments detail
        ws.cell(rr, 1, "Fertilisers-separation equity-bridge adjustments (methodology §4.2)"); ws.cell(rr, 1).font = SUB; rr += 1
        hdr = rr
        for cc, t in ((1, "Item"), (2, "Amount"), (3, "Treatment"), (4, "Net effect")):
            ws.cell(hdr, cc, t).font = BOLD
        rr += 1
        nb = inp["company_raw"]["normalised_baseline"]
        adj_cells = []
        for a in nb["equity_bridge_adjustments"]:
            ws.cell(rr, 1, a.get("description", a["id"])[:70])
            camt = ws.cell(rr, 2, a["amount_aud_m"]); camt.fill = YELLOW; camt.font = BLUEFONT; camt.number_format = NUM1
            treat = a.get("treatment"); ws.cell(rr, 3, treat)
            sign = 1.0 if a["direction"] == "subtract_from_equity" else -1.0
            if treat == "add_back_in_full":
                netf = f"={sign}*B{rr}"
            elif treat == "add_back_gap_only":
                prov = a.get("provided_for_at_anchor_aud_m", 0.0)
                cprov = ws.cell(rr, 5, prov); cprov.fill = YELLOW; cprov.font = BLUEFONT; cprov.number_format = NUM1
                netf = f"={sign}*(B{rr}-E{rr})"
            elif treat == "probability_weighted":
                prob = a["probability"]
                cprob = ws.cell(rr, 6, prob); cprob.fill = YELLOW; cprob.font = BLUEFONT; cprob.number_format = NUM1 + "00"
                netf = f"={sign}*B{rr}*F{rr}"
            else:
                netf = "=0"
            cn = ws.cell(rr, 4, netf); cn.number_format = NUM1
            adj_cells.append(f"D{rr}"); rr += 1
        ws.cell(rr, 1, "Net equity-bridge adjustments"); ws.cell(rr, 1).font = BOLD
        c = ws.cell(rr, 2, "=" + "+".join(adj_cells)); c.number_format = NUM1; c.font = BOLD
        self.adj_row = rr; self.ref["adj_net"] = f"'Equity bridge'!$B${rr}"; rr += 1
        rr += 1
        self._eq_ws = ws
        self._eq_rr = rr
        self._eq_b11 = b11
        ws.column_dimensions["A"].width = 52
        for cc in "BCDEF":
            ws.column_dimensions[cc].width = 14

    def finish_equity(self):
        R = self.ref
        ws = self._eq_ws
        ws.cell(self._eq_rr, 1, "Per-share bridge (central case)"); ws.cell(self._eq_rr, 1).font = SUB; self._eq_rr += 1
        ws.cell(self._eq_rr, 1, "Enterprise value (from DCF, Muddle Through)"); ws.cell(self._eq_rr, 2, f"={R['ev:muddle_through']}").number_format = NUM1; ev_r = self._eq_rr; self._eq_rr += 1
        ws.cell(self._eq_rr, 1, "less: net debt at valuation"); ws.cell(self._eq_rr, 2, f"=-B{self._eq_b11}").number_format = NUM1; nd_r = self._eq_rr; self._eq_rr += 1
        ws.cell(self._eq_rr, 1, "less: equity-bridge adjustments (net)"); ws.cell(self._eq_rr, 2, f"=-B{self.adj_row}").number_format = NUM1; adj_r = self._eq_rr; self._eq_rr += 1
        ws.cell(self._eq_rr, 1, "less: AASB 16 lease liabilities"); ws.cell(self._eq_rr, 2, f"=-{R['leases']}").number_format = NUM1; le_r = self._eq_rr; self._eq_rr += 1
        ws.cell(self._eq_rr, 1, "Equity value"); ws.cell(self._eq_rr, 1).font = BOLD
        ce = ws.cell(self._eq_rr, 2, f"=B{ev_r}+B{nd_r}+B{adj_r}+B{le_r}"); ce.number_format = NUM1; ce.font = BOLD; eq_r = self._eq_rr; self._eq_rr += 1
        ws.cell(self._eq_rr, 1, "Value per share"); ws.cell(self._eq_rr, 1).font = BOLD
        cv = ws.cell(self._eq_rr, 2, f"=B{eq_r}/{R['shares']}"); cv.number_format = NUM1 + "00"; cv.font = BOLD; self._eq_rr += 1
        ws.cell(self._eq_rr, 1, "Discount / (premium) vs market"); ws.cell(self._eq_rr, 2, f"=B{self._eq_rr-1}/{R['mkt']}-1").number_format = PCT2

        ws.column_dimensions["A"].width = 52
        for cc in "BCDEF":
            ws.column_dimensions[cc].width = 14

    # ---- Scenarios summary -------------------------------------------------
    def build_scenarios(self):
        ws = self.wb.create_sheet("Scenarios")
        R = self.ref
        ws.cell(1, 1, "Scenario summary — value per share by world (each links to DCF build)"); ws.cell(1, 1).font = HDR
        hr = 3
        for cc, t in ((1, "Scenario"), (2, "Value/share"), (3, "EV (AUD m)"), (4, "vs market")):
            ws.cell(hr, cc, t).font = BOLD
        rr = hr + 1
        for sid, nm in SCEN:
            ws.cell(rr, 1, nm)
            ws.cell(rr, 2, f"={R[f'vps:{sid}']}").number_format = NUM1 + "00"
            ws.cell(rr, 3, f"={R[f'ev:{sid}']}").number_format = NUM0
            ws.cell(rr, 4, f"={R[f'vps:{sid}']}/{R['mkt']}-1").number_format = PCT2
            rr += 1
        rr += 1
        ws.cell(rr, 1, "Market reference price"); ws.cell(rr, 2, f"={R['mkt']}").number_format = NUM1 + "0"
        ws.column_dimensions["A"].width = 28
        for cc in "BCD":
            ws.column_dimensions[cc].width = 14

    # ---- Comparables & beta (triangulation) --------------------------------
    def build_comps(self):
        import beta_data as _bd
        d = _bd.DNL
        ws = self.wb.create_sheet("Comparables & beta")
        R = self.ref
        ws.cell(1, 1, "Beta triangulation — peer unlever/relever (MOCK peer data, pending EODHD feed)"); ws.cell(1, 1).font = HDR
        ws.cell(2, 1, "Asset beta = equity beta / (1+(1-tax)*D/E); relevered to DNL gearing & tax. Ratified beta is the owner's triangulated choice, not a mechanical median."); ws.cell(2, 1).font = NOTE
        # DNL relever basis
        rr = 4
        ws.cell(rr, 1, "DNL D/E (for relever)"); c = ws.cell(rr, 2, d["subject"]["de"]); c.fill = YELLOW; c.font = BLUEFONT; c.number_format = NUM1 + "00"; dnl_de = rr; rr += 1
        ws.cell(rr, 1, "DNL tax (for relever)"); ws.cell(rr, 2, f"={R['tax_wacc']}").number_format = PCT2; dnl_tax = rr; rr += 1
        rr += 1
        hdr = rr
        for cc, t in ((1, "Peer"), (2, "Equity beta"), (3, "Tax"), (4, "D/E"), (5, "Include?"), (6, "Asset beta"), (7, "Relevered to DNL")):
            ws.cell(hdr, cc, t).font = BOLD
        rr += 1
        relev_cells = []
        idx = d["indexDefault"]; win = d["windowDefault"]
        for comp in d["comparables"]:
            be = comp["data"][idx][win]["beta"]
            ws.cell(rr, 1, comp["name"])
            cb = ws.cell(rr, 2, be); cb.fill = YELLOW; cb.font = BLUEFONT; cb.number_format = NUM1 + "00"
            ct = ws.cell(rr, 3, comp["tax"]); ct.fill = YELLOW; ct.font = BLUEFONT; ct.number_format = PCT2
            cd = ws.cell(rr, 4, comp["gearingDE"]); cd.fill = YELLOW; cd.font = BLUEFONT; cd.number_format = NUM1 + "00"
            inc = "Yes" if comp.get("selected") else "No (outlier)"
            ci = ws.cell(rr, 5, inc); ci.fill = YELLOW; ci.font = BLUEFONT
            ws.cell(rr, 6, f"=B{rr}/(1+(1-C{rr})*D{rr})").number_format = NUM1 + "00"
            ws.cell(rr, 7, f"=F{rr}*(1+(1-B{dnl_tax})*B{dnl_de})").number_format = NUM1 + "00"
            if comp.get("selected"):
                relev_cells.append(f"G{rr}")
            rr += 1
        rr += 1
        ws.cell(rr, 1, "Median relevered beta (included peers)"); ws.cell(rr, 1).font = BOLD
        ws.cell(rr, 2, "=MEDIAN(" + ",".join(relev_cells) + ")").number_format = NUM1 + "00"; ws.cell(rr, 2).font = BOLD; rr += 1
        ws.cell(rr, 1, "Ratified beta (used in WACC, from Assumptions)"); ws.cell(rr, 1).font = BOLD
        cr = ws.cell(rr, 2, f"={R['beta']}"); cr.number_format = NUM1 + "00"; cr.font = BOLD; rr += 1
        ws.cell(rr, 1, d["subject"].get("measuredNote", "")); ws.cell(rr, 1).font = NOTE
        ws.column_dimensions["A"].width = 40
        for cc in "BCDEFG":
            ws.column_dimensions[cc].width = 15

    # ---- Porter's five forces (feeds the revenue-chain offset) -------------
    def build_porters(self):
        ws = self.wb.create_sheet("Porter five forces")
        R = self.ref
        ws.cell(1, 1, "Porter's five forces \u2192 company-position offset (feeds Revenue growth B41)"); ws.cell(1, 1).font = HDR
        ws.cell(2, 1, "Each assessed sub-offset is a yellow input on Assumptions; they sum to the net company-position offset used in the revenue chain."); ws.cell(2, 1).font = NOTE
        rr = 4
        for cc, t in ((1, "Force / channel"), (2, "Assessed offset")):
            ws.cell(rr, 1 if cc == 1 else 2, t).font = BOLD
        rr += 1
        rows = [
            ("Rivalry \u2014 competitive position", R["ff_riv"]),
            ("Rivalry \u2014 product mix", R["ff_mix"]),
            ("New entrants \u2014 pipeline uplift", R["ff_ent"]),
            ("Buyer / supplier / substitutes (net)", R["ff_bss"]),
        ]
        cells = []
        for label, ref in rows:
            ws.cell(rr, 1, label); ws.cell(rr, 2, f"={ref}").number_format = PCT3; cells.append(f"B{rr}"); rr += 1
        ws.cell(rr, 1, "Net company-position offset"); ws.cell(rr, 1).font = BOLD
        ws.cell(rr, 2, "=" + "+".join(cells)).number_format = PCT3; ws.cell(rr, 2).font = BOLD; net_r = rr; rr += 1
        rr += 1
        ws.cell(rr, 1, "Ties Revenue growth B41 (net offset)"); ws.cell(rr, 2, f"={R['rev_growth:muddle_through']}*0+'Revenue growth'!C" + str(self.rev_rows["B41"])).number_format = PCT3
        ws.cell(rr, 3, "\u2190 same value, cross-checked").font = NOTE
        ws.column_dimensions["A"].width = 40; ws.column_dimensions["B"].width = 16; ws.column_dimensions["C"].width = 26

    # ---- Comparability metrics (drivers behind the beta triangulation) -----
    def build_comparability(self, beta):
        ws = self.wb.create_sheet("Comparability metrics")
        ws.cell(1, 1, "Comparability drivers — what makes each peer more or less like DNL (MOCK, pending feed)"); ws.cell(1, 1).font = HDR
        ws.cell(2, 1, "Read across the three drivers to judge asset-beta comparability; higher = more exposed. Financial leverage regears beta; operating leverage and cyclicality shape the asset beta."); ws.cell(2, 1).font = NOTE
        hr = 4
        cols = [("Name", 1), ("Financial leverage (net debt/EBITDA)", 2), ("Operating leverage (DOL)", 3),
                ("Revenue cyclicality (0-1)", 4), ("Gearing D/E", 5), ("Tax", 6), ("Role", 7)]
        for t, c in cols:
            cc = ws.cell(hr, c, t); cc.font = BOLD; cc.alignment = Alignment(wrap_text=True, vertical="top")
        rr = hr + 1
        subj = beta["subject"]; sdet = subj["det"]
        ws.cell(rr, 1, subj["name"] + " (subject)").font = BOLD
        for c, v, fmt in ((2, sdet["ndeb"], NUM1 + "0"), (3, sdet["dol"], NUM1 + "0"), (4, sdet["cyc"], NUM1 + "00"),
                          (5, subj["de"], NUM1 + "00"), (6, subj["tax"], PCT2)):
            cell = ws.cell(rr, c, v); cell.fill = YELLOW; cell.font = BLUEFONT; cell.number_format = fmt
        ws.cell(rr, 7, "Subject")
        rr += 1
        subj_rows = []
        for comp in beta["comparables"]:
            det = comp.get("det") or {}
            ws.cell(rr, 1, comp["name"])
            for c, v, fmt in ((2, det.get("ndeb"), NUM1 + "0"), (3, det.get("dol"), NUM1 + "0"), (4, det.get("cyc"), NUM1 + "00"),
                              (5, comp.get("gearingDE"), NUM1 + "00"), (6, comp.get("tax"), PCT2)):
                cell = ws.cell(rr, c, v); cell.fill = YELLOW; cell.font = BLUEFONT; cell.number_format = fmt
            ws.cell(rr, 7, "Peer" if comp.get("selected") else "Peer (outlier)")
            subj_rows.append(rr); rr += 1
        rr += 1
        ws.cell(rr, 1, "Peer average (all listed)").font = BOLD
        for c, fmt in ((2, NUM1 + "0"), (3, NUM1 + "0"), (4, NUM1 + "00"), (5, NUM1 + "00"), (6, PCT2)):
            col = _col(c)
            ws.cell(rr, c, f"=AVERAGE({col}{subj_rows[0]}:{col}{subj_rows[-1]})").number_format = fmt
        rr += 2
        ws.cell(rr, 1, beta.get("detNote", "").replace("<b>", "").replace("</b>", "").replace("&rsquo;", "'").replace("&mdash;", "—")[:600]); ws.cell(rr, 1).font = NOTE
        ws.column_dimensions["A"].width = 26
        for c in "BCDEFG":
            ws.column_dimensions[c].width = 16

    # ---- Trading multiples (peers + DNL implied + market-implied) ----------
    def build_multiples(self, beta, cfg):
        ws = self.wb.create_sheet("Trading multiples")
        R = self.ref
        MULT = '0.0"x"'
        ws.cell(1, 1, "Trading multiples — peer comps, DNL implied value, and the market-implied read"); ws.cell(1, 1).font = HDR
        ws.cell(2, 1, "Peer EV = market cap + net debt; multiples computed by formula. DNL implied value applies the peer median to DNL's own earnings, then bridges to equity."); ws.cell(2, 1).font = NOTE
        # DNL net debt incl leases (lease-inclusive), from Assumptions run-rates
        ndincl = f"({R['nd_anchor']}-{R['ocf_rr']}*{R['pa_days']}/365+{R['capex_rr']}*{R['pa_days']}/365+{R['leases']})"

        # ---- peer table ----
        hr = 4
        heads = ["Peer", "Price", "Shares (m)", "Net debt", "Market cap", "EV",
                 "EV/EBITDA (ttm)", "EV/EBITDA (fwd)", "EV/EBIT (ttm)", "P/E (ttm)", "P/E (fwd)"]
        for j, t in enumerate(heads):
            cc = ws.cell(hr, 1 + j, t); cc.font = BOLD; cc.alignment = Alignment(wrap_text=True, vertical="top")
        rr = hr + 1
        peer_rows = []
        for comp in beta["comparables"]:
            m = comp["mfin"]
            ws.cell(rr, 1, comp["name"])
            for c, v, fmt in ((2, m["price"], NUM1 + "0"), (3, m["shares"], NUM0), (4, m["netDebt"], NUM0)):
                cell = ws.cell(rr, c, v); cell.fill = YELLOW; cell.font = BLUEFONT; cell.number_format = fmt
            # hidden yellow earnings to the right (cols 12-17): EBITDA ttm/fwd, EBIT ttm, NI ttm/fwd
            for c, v in ((12, m["ebitda"]["ttm"]), (13, m["ebitda"]["fwd"]), (14, m["ebit"]["ttm"]),
                         (15, m["ni"]["ttm"]), (16, m["ni"]["fwd"])):
                cell = ws.cell(rr, c, v); cell.fill = YELLOW; cell.font = BLUEFONT; cell.number_format = NUM0
            ws.cell(rr, 5, f"=B{rr}*C{rr}").number_format = NUM0          # market cap
            ws.cell(rr, 6, f"=E{rr}+D{rr}").number_format = NUM0          # EV
            ws.cell(rr, 7, f"=F{rr}/L{rr}").number_format = MULT          # EV/EBITDA ttm
            ws.cell(rr, 8, f"=F{rr}/M{rr}").number_format = MULT          # EV/EBITDA fwd
            ws.cell(rr, 9, f"=F{rr}/N{rr}").number_format = MULT          # EV/EBIT ttm
            ws.cell(rr, 10, f"=E{rr}/O{rr}").number_format = MULT         # P/E ttm
            ws.cell(rr, 11, f"=E{rr}/P{rr}").number_format = MULT         # P/E fwd
            peer_rows.append(rr); rr += 1
        # median row
        med_r = rr
        ws.cell(rr, 1, "Peer median").font = BOLD
        for c in range(7, 12):
            col = _col(c)
            cell = ws.cell(rr, c, f"=MEDIAN({col}{peer_rows[0]}:{col}{peer_rows[-1]})"); cell.number_format = MULT; cell.font = BOLD
        self._mult_med = {"evebitda_ttm": f"G{med_r}", "evebitda_fwd": f"H{med_r}", "evebit_ttm": f"I{med_r}",
                          "pe_ttm": f"J{med_r}", "pe_fwd": f"K{med_r}"}
        rr += 2

        # ---- DNL earnings bases ----
        ws.cell(rr, 1, "DNL earnings bases (our build; consensus where shown)"); ws.cell(rr, 1).font = SUB; rr += 1
        bh = rr
        for j, t in enumerate(["Base", "EBITDA", "EBIT", "Net income", "EBITDA (consensus)", "EBIT (cons.)", "NI (cons.)"]):
            ws.cell(bh, 1 + j, t).font = BOLD
        rr += 1
        bases = cfg["multiples"]["bases"]
        base_order = [k for k in ("fy25u", "fy26", "fy27") if k in bases]
        base_rows = {}
        for k in base_order:
            b = bases[k]
            ws.cell(rr, 1, b.get("label", k))
            for c, v in ((2, b["ebitda"]), (3, b["ebit"]), (4, b["ni"])):
                cell = ws.cell(rr, c, v); cell.fill = YELLOW; cell.font = BLUEFONT; cell.number_format = NUM0
            cons = b.get("consensus")
            if cons:
                for c, v in ((5, cons["ebitda"]), (6, cons["ebit"]), (7, cons["ni"])):
                    cell = ws.cell(rr, c, v); cell.fill = YELLOW; cell.font = BLUEFONT; cell.number_format = NUM0
            base_rows[k] = rr; rr += 1
        rr += 1

        # ---- DNL implied value per share (apply peer median to DNL earnings) ----
        ws.cell(rr, 1, "DNL implied value per share (peer median x DNL earnings, bridged to equity)"); ws.cell(rr, 1).font = SUB; rr += 1
        ih = rr
        for j, t in enumerate(["Base", "via EV/EBITDA", "via EV/EBIT", "via P/E"]):
            ws.cell(ih, 1 + j, t).font = BOLD
        rr += 1
        # use fwd multiples for forward bases, ttm for the trailing FY25u
        for k in base_order:
            br = base_rows[k]
            fwd = (k != "fy25u")
            evebitda = self._mult_med["evebitda_fwd"] if fwd else self._mult_med["evebitda_ttm"]
            pe = self._mult_med["pe_fwd"] if fwd else self._mult_med["pe_ttm"]
            evebit = self._mult_med["evebit_ttm"]
            ws.cell(rr, 1, bases[k].get("label", k))
            ws.cell(rr, 2, f"=({evebitda}*B{br}-{ndincl})/{R['shares']}").number_format = NUM1 + "00"
            ws.cell(rr, 3, f"=({evebit}*C{br}-{ndincl})/{R['shares']}").number_format = NUM1 + "00"
            ws.cell(rr, 4, f"={pe}*D{br}/{R['shares']}").number_format = NUM1 + "00"
            rr += 1
        rr += 1

        # ---- market-implied read ----
        ws.cell(rr, 1, "Market-implied read (what the tape pays on DNL's own earnings)"); ws.cell(rr, 1).font = SUB; rr += 1
        fy26 = base_rows.get("fy26", base_rows[base_order[-1]])
        ws.cell(rr, 1, "DNL market cap = price x shares"); ws.cell(rr, 2, f"={R['mkt']}*{R['shares']}").number_format = NUM0; mc = rr; rr += 1
        ws.cell(rr, 1, "DNL EV = market cap + net debt (incl leases)"); ws.cell(rr, 2, f"=B{mc}+{ndincl}").number_format = NUM0; ev = rr; rr += 1
        ws.cell(rr, 1, "Market EV/EBITDA (on FY26 EBITDA)"); ws.cell(rr, 2, f"=B{ev}/B{fy26}").number_format = MULT; rr += 1
        ws.cell(rr, 1, "Market EV/EBIT (on FY26 EBIT)"); ws.cell(rr, 2, f"=B{ev}/C{fy26}").number_format = MULT; rr += 1
        ws.cell(rr, 1, "Market P/E (on FY26 NI)"); ws.cell(rr, 2, f"=B{mc}/D{fy26}").number_format = MULT; rr += 1
        rr += 1
        ws.cell(rr, 1, "Cross-check: DCF value per share (Muddle Through)").font = BOLD
        ws.cell(rr, 2, f"={R['vps:muddle_through']}").number_format = NUM1 + "00"; ws.cell(rr, 2).font = BOLD; rr += 1
        ws.cell(rr, 1, cfg["multiples"].get("note", "")[:300]); ws.cell(rr, 1).font = NOTE

        for c in "A":
            ws.column_dimensions[c].width = 42
        for c in "BCDEFGHIJK":
            ws.column_dimensions[c].width = 13
        for c in ("L", "M", "N", "O", "P"):
            ws.column_dimensions[c].width = 11

    # ---- AASB 16 lease detail ---------------------------------------------
    def build_lease(self, cfg):
        ws = self.wb.create_sheet("Lease detail")
        R = self.ref
        lc = cfg.get("_leaseContract", {})
        ws.cell(1, 1, "AASB 16 lease detail (treated as debt, Approach A) — MOCK shape pending EODHD feed"); ws.cell(1, 1).font = HDR
        rr = 3
        ws.cell(rr, 1, "Lease liability (from Assumptions)"); ws.cell(rr, 2, f"={R['leases']}").number_format = NUM1; rr += 1
        ws.cell(rr, 1, "Annual lease cost (RoU dep + interest)"); c = ws.cell(rr, 2, lc.get("annualLeaseCost")); c.fill = YELLOW; c.font = BLUEFONT; c.number_format = NUM1; alc = rr; rr += 1
        ws.cell(rr, 1, "Incremental borrowing rate (IBR)"); c = ws.cell(rr, 2, lc.get("incrementalBorrowingRate")); c.fill = YELLOW; c.font = BLUEFONT; c.number_format = PCT2; ibr = rr; rr += 1
        rr += 1
        ws.cell(rr, 1, "Undiscounted lease maturity"); ws.cell(rr, 1).font = SUB; rr += 1
        mh = rr
        ws.cell(mh, 1, "Year").font = BOLD; ws.cell(mh, 2, "Undisc. payment").font = BOLD; ws.cell(mh, 3, "Discount factor").font = BOLD; ws.cell(mh, 4, "PV at IBR").font = BOLD
        rr += 1
        mat = lc.get("leaseMaturityUndisc", {})
        seq = [("y1", 1), ("y2", 2), ("y3", 3), ("y4", 4), ("y5", 5), ("beyond5", 7)]
        pv_cells = []; pay_cells = []
        for key, t in seq:
            if key not in mat:
                continue
            ws.cell(rr, 1, "Beyond 5 (mid-pt)" if key == "beyond5" else f"Year {t}")
            cp = ws.cell(rr, 2, mat[key]); cp.fill = YELLOW; cp.font = BLUEFONT; cp.number_format = NUM1
            ws.cell(rr, 3, f"=1/(1+B{ibr})^{t}").number_format = NUM1 + "000"
            ws.cell(rr, 4, f"=B{rr}*C{rr}").number_format = NUM1
            pv_cells.append(f"D{rr}"); pay_cells.append(f"B{rr}"); rr += 1
        ws.cell(rr, 1, "Total undiscounted").font = BOLD
        ws.cell(rr, 2, "=" + "+".join(pay_cells)).number_format = NUM1; ws.cell(rr, 2).font = BOLD; rr += 1
        ws.cell(rr, 1, "PV of lease payments (check vs liability)").font = BOLD
        ws.cell(rr, 2, "=" + "+".join(pv_cells)).number_format = NUM1; ws.cell(rr, 2).font = BOLD; rr += 2
        ws.cell(rr, 1, "Lease-neutral view (EBITDAR)"); ws.cell(rr, 1).font = SUB; rr += 1
        ws.cell(rr, 1, "Add back annual lease cost to EBITDA (peer-uniform comparison)")
        ws.cell(rr, 2, f"=B{alc}").number_format = NUM1; rr += 1
        ws.cell(rr, 1, lc.get("contractNote", "")[:400]); ws.cell(rr, 1).font = NOTE
        ws.column_dimensions["A"].width = 46
        for c in "BCD":
            ws.column_dimensions[c].width = 15

    def to_bytes(self):
        buf = io.BytesIO()
        self.wb.save(buf)
        return buf.getvalue()


def build_dnl_workbook_bytes(cfg=None):
    """cfg is the assembled ``dnl`` config dict (has ``multiples``, ``_leaseContract``,
    ``beta``). When called from build_cfgs it is passed in; for a standalone run it is
    read from cfgs_gen.json so the multiples / lease / comparability sheets can build."""
    import json as _json
    import beta_data as _bd
    if cfg is None:
        _cfp = _ROOT / "ui_prototypes" / "_generator" / "cfgs_gen.json"
        cfg = _json.load(open(_cfp))["dnl"] if _cfp.exists() else {}
    beta = _bd.DNL
    inp = _load_central()
    craw = inp["company_raw"]; nb = craw["normalised_baseline"]
    coff = nb["revenue_growth_chain"]["shared"]["company_offset"]
    b = Book()
    # DM/EM region partition for the revenue sheet
    dm = set(coff["developed_market_regions"])
    geo = _geographic_regions(craw)
    b._dm_regions = [g["geo"] for g in geo if g["geo"] in dm]
    b._em_regions = [g["geo"] for g in geo if g["geo"] not in dm]
    b.build_assumptions(inp)
    b.build_revenue(inp)
    b.build_wacc()
    b.build_tax()
    b.build_equity(inp)   # net-debt walk + adjustments; defines adj_row used by DCF
    b.build_dcf()         # six-scenario DCF; defines EV/vps refs
    b.finish_equity()     # per-share bridge (central case) links DCF EV
    b.build_comps()
    b.build_comparability(beta)
    b.build_porters()
    if cfg.get("multiples"):
        b.build_multiples(beta, cfg)
    if cfg.get("_leaseContract"):
        b.build_lease(cfg)
    b.build_scenarios()
    return b.to_bytes()


# ======================================================================
# WBC bank workbook (methodology §15: residual-income / DDM, no EV bridge)
# ======================================================================

def build_wbc_workbook_bytes(cfg=None):
    """Full audited WBC bank workbook, formula-linked, tying the bank engine (MT 30.03)."""
    import beta_data as _bd
    from vcc_valuations.translator import (
        load_inputs as _li, resolve_normalised_baseline as _rnb,
        build_bank_inputs_from_data as _bbi,
    )
    from vcc_valuations.dcf.bank_engine import BankEngine as _BE

    if cfg is None:
        import json as _json
        _cfp = _ROOT / "ui_prototypes" / "_generator" / "cfgs_gen.json"
        cfg = _json.load(open(_cfp))["wbc"] if _cfp.exists() else {}
    inp = _li(_ROOT, "muddle_through", "australian_major_banks", "wbc")
    nb = inp["company_raw"]["normalised_baseline"]
    bb = nb["bank_build"]
    coe = _rnb(inp)["cost_of_equity_build"]
    chain = bb["aiea_nim_chain"]; shared = chain["shared"]
    ov = bb["overlays"]; base = ov["baseline"]
    anchors = bb["balance_sheet_anchors"]; drv = bb["forecast_drivers"]
    income = bb["income_anchors_1h26"]; timing = bb["timing"]

    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    R = {}
    MULT = '0.0"x"'

    # ---------------- Assumptions ----------------
    ws = wb.active; ws.title = "Assumptions"; ws.sheet_properties.tabColor = "FFD966"
    r = [1]
    def head(t): ws.cell(r[0], 1, t).font = HDR; r[0] += 1
    def sub(t): ws.cell(r[0], 1, t).font = SUB; r[0] += 1
    def note(t): ws.cell(r[0], 1, t).font = NOTE; r[0] += 1
    def line(label, value, key, fmt=None):
        ws.cell(r[0], 1, label)
        c = ws.cell(r[0], 2, value); c.fill = YELLOW; c.font = BLUEFONT
        if fmt: c.number_format = fmt
        R[key] = f"'Assumptions'!$B${r[0]}"; r[0] += 1
    head("WBC — Assumptions (yellow cells are the only inputs; bank archetype §15)")
    note("Every other sheet links here by formula. Sourced from the production bank engine's data.")
    r[0] += 1
    sub("Cost of equity (single Ke, held across scenarios)")
    line("Risk-free rate Rf", coe["risk_free_rate"], "rf", PCT2)
    line("Equity risk premium ERP", coe["equity_risk_premium"], "erp", PCT2)
    line("Beta (peer-triangulated)", coe["beta"], "beta", NUM1 + "0")
    r[0] += 1
    sub("AIEA / NIM chain (industry anchor + company Five-Forces offset; standing rule 1)")
    line("Industry AIEA growth", shared["industry"]["aiea_growth"], "ind_aiea", PCT2)
    line("Industry NIM anchor", shared["industry"]["nim_anchor"], "ind_nim", PCT3)
    line("WBC AIEA growth offset", shared["company_offset"]["aiea_growth_offset"], "off_aiea", PCT2)
    line("WBC NIM offset", shared["company_offset"]["nim_offset"], "off_nim", PCT3)
    r[0] += 1
    sub("1H26 income anchors (AUD m, half-year)")
    line("Non-interest income (1H26)", income["non_interest_income"], "nonint_1h", NUM0)
    r[0] += 1
    sub("Balance-sheet anchors (AUD m, 1H26 close)")
    line("Average interest-earning assets", anchors["aiea_1h26_average"], "aiea", NUM0)
    line("Book equity", anchors["book_equity"], "book_eq", NUM0)
    line("AT1 hybrid", anchors["at1_hybrid"], "at1", NUM0)
    line("Non-controlling interests", anchors["non_controlling_interests"], "nci", NUM0)
    line("Treasury shares", anchors["treasury_shares"], "treas", NUM0)
    line("Shares outstanding (m)", anchors["shares_outstanding_m"], "shares", NUM1)
    r[0] += 1
    sub("Forecast drivers")
    line("Non-interest income growth", drv["non_interest_income_growth"], "ni_g", PCT2)
    line("Average loans as % of AIEA", drv["avg_loans_pct_aiea"], "loans_pct", PCT2)
    line("Effective tax rate", drv["effective_tax_rate"], "tax", PCT2)
    line("Dividend payout ratio", drv["dividend_payout_ratio"], "payout", PCT2)
    r[0] += 1
    sub("Timing")
    line("Stub years", timing["stub_years"], "stub", NUM1 + "00")
    line("Horizon years", timing["horizon_years"], "horizon", "0")
    line("AIEA Y1 time factor", timing["aiea_y1_time_factor"], "aiea_y1f", NUM1 + "00")
    r[0] += 1
    # per-scenario macro deltas + overlay endpoints (scenarios as columns C..H)
    sub("Per-scenario drivers")
    hdr = r[0]; ws.cell(hdr, 1, "Driver").font = BOLD
    for j, (sid, nm) in enumerate(SCEN):
        cc = ws.cell(hdr, 3 + j, nm); cc.font = BOLD; cc.alignment = Alignment(horizontal="right", wrap_text=True)
    r[0] += 1
    def scen_row(label, key, fn, fmt):
        ws.cell(r[0], 1, label)
        for j, (sid, nm) in enumerate(SCEN):
            c = ws.cell(r[0], 3 + j, fn(sid)); c.fill = YELLOW; c.font = BLUEFONT; c.number_format = fmt
            R[f"{key}:{sid}"] = f"'Assumptions'!${_col(3 + j)}${r[0]}"
        r[0] += 1
    scen_row("AIEA growth delta", "d_aiea", lambda s: chain["by_scenario"][s]["aiea_growth_delta"], PCT2)
    scen_row("NIM delta (bps)", "d_nimbps", lambda s: chain["by_scenario"][s]["nim_delta_bps"], "0")
    scen_row("Y5 cost-to-income", "y5cti", lambda s: ov["by_scenario"][s]["y5_cost_to_income"], PCT2)
    scen_row("Credit-loss delta (bps)", "d_clbps", lambda s: ov["by_scenario"][s]["credit_loss_delta_bps"], "0")
    scen_row("Terminal ROE", "troe", lambda s: ov["by_scenario"][s]["terminal_roe"], PCT2)
    scen_row("Terminal growth g", "tg", lambda s: ov["by_scenario"][s]["terminal_growth"], PCT2)
    r[0] += 1
    # baseline per-year glides (stub..Y5 as columns C..H)
    sub("Baseline per-year glides (Muddle Through): [stub, Y1..Y5]")
    ph = r[0]; ws.cell(ph, 1, "Line").font = BOLD
    for j, lab in enumerate(["Stub", "Y1", "Y2", "Y3", "Y4", "Y5"]):
        ws.cell(ph, 3 + j, lab).font = BOLD
    r[0] += 1
    def glide_row(label, key, vec, fmt):
        ws.cell(r[0], 1, label)
        for j, v in enumerate(vec):
            c = ws.cell(r[0], 3 + j, v); c.fill = YELLOW; c.font = BLUEFONT; c.number_format = fmt
            R[f"{key}:{j}"] = f"'Assumptions'!${_col(3 + j)}${r[0]}"
        r[0] += 1
    glide_row("NIM applied", "nim", base["nim_applied"], PCT3)
    glide_row("Cost-to-income", "cti", base["cost_to_income"], PCT2)
    glide_row("Credit-loss rate", "clr", base["credit_loss_rate"], PCT3)
    ws.column_dimensions["A"].width = 40
    for cc in "BCDEFGH": ws.column_dimensions[cc].width = 13

    # ---------------- Ke build ----------------
    ws = wb.create_sheet("Ke build")
    ws.cell(1, 1, "Cost of equity (bank archetype — no WACC / EV bridge)").font = HDR
    ws.cell(3, 1, "Re = Rf + beta x ERP"); c = ws.cell(3, 2, f"={R['rf']}+{R['beta']}*{R['erp']}"); c.number_format = PCT3; c.font = BOLD
    R["ke"] = "'Ke build'!$B$3"
    ws.column_dimensions["A"].width = 30

    # ---------------- AIEA/NIM chain ----------------
    ws = wb.create_sheet("AIEA-NIM chain")
    ws.cell(1, 1, "AIEA growth & NIM anchor by scenario (industry + company offset + scenario)").font = HDR
    ws.cell(2, 1, "Standing rule 1: industry anchor and company offset are separate inputs; the applied value is derived.").font = NOTE
    hr = 4; ws.cell(hr, 1, "Derived").font = BOLD
    for j, (sid, nm) in enumerate(SCEN):
        ws.cell(hr, 3 + j, nm).font = BOLD
    ws.cell(hr + 1, 1, "WBC AIEA growth")
    ws.cell(hr + 2, 1, "WBC NIM anchor")
    for j, (sid, nm) in enumerate(SCEN):
        col = _col(3 + j)
        ws.cell(hr + 1, 3 + j, f"={R['ind_aiea']}+{R[f'd_aiea:{sid}']}+{R['off_aiea']}").number_format = PCT3
        ws.cell(hr + 2, 3 + j, f"={R['ind_nim']}+{R[f'd_nimbps:{sid}']}/10000+{R['off_nim']}").number_format = PCT3
        R[f"aieag:{sid}"] = f"'AIEA-NIM chain'!{col}{hr+1}"
    ws.column_dimensions["A"].width = 22
    for cc in "CDEFGH": ws.column_dimensions[cc].width = 15

    # ---------------- P&L Forecast (six scenarios as columns) ----------------
    ws = wb.create_sheet("P&L Forecast")
    ws.cell(1, 1, "P&L forecast — six scenarios (AUD m). NII = AIEA x NIM x period; §15 operating build.").font = HDR
    hr = 3; ws.cell(hr, 1, "Line").font = BOLD
    SIDS = [sid for sid, _ in SCEN]; SC = [_col(3 + j) for j in range(6)]
    for j, (sid, nm) in enumerate(SCEN):
        ws.cell(hr, 3 + j, nm).font = BOLD
    rr = [hr + 1]; rmap = {}
    PL = ["Stub", "Y1", "Y2", "Y3", "Y4", "Y5"]
    def band(t): ws.cell(rr[0], 1, t).font = SUB; rr[0] += 1
    def prow(key, label, fmt, fn, bold=False):
        ws.cell(rr[0], 1, label)
        if bold: ws.cell(rr[0], 1).font = BOLD
        for j, sid in enumerate(SIDS):
            c = ws.cell(rr[0], 3 + j, fn(sid, SC[j], j)); c.number_format = fmt
            if bold: c.font = BOLD
        rmap[key] = rr[0]; rr[0] += 1
    # period length memo
    prow("plen", "Period length (yrs)", NUM1 + "00",
         lambda s, c, j, : f"={R['stub']}") if False else None
    band("AIEA (average interest-earning assets)")
    for p in range(6):
        def f_aiea(s, c, j, p=p):
            if p == 0: return f"={R['aiea']}*(1+{R[f'aieag:{s}']}*{R['stub']}/2)"
            if p == 1: return f"={R['aiea']}*(1+{R[f'aieag:{s}']}*{R['aiea_y1f']})"
            return f"={c}{rmap[f'aiea{p-1}']}*(1+{R[f'aieag:{s}']})"
        prow(f"aiea{p}", f"  {PL[p]} AIEA", NUM0, f_aiea)
    band("NIM applied")
    for p in range(6):
        prow(f"nim{p}", f"  {PL[p]} NIM", PCT3,
             lambda s, c, j, p=p: f"={R[f'nim:{p}']}+{R[f'd_nimbps:{s}']}/10000")
    band("Net interest income")
    for p in range(6):
        plen = f"{R['stub']}" if p == 0 else "1"
        prow(f"nii{p}", f"  {PL[p]} NII", NUM1,
             lambda s, c, j, p=p, plen=plen: f"={c}{rmap[f'aiea{p}']}*{c}{rmap[f'nim{p}']}*{plen}")
    band("Non-interest income")
    for p in range(6):
        def f_ni(s, c, j, p=p):
            if p == 0: return f"={R['nonint_1h']}*{R['stub']}*2"
            if p == 1: return f"=2*{R['nonint_1h']}*(1+{R['ni_g']})"
            return f"={c}{rmap[f'ni{p-1}']}*(1+{R['ni_g']})"
        prow(f"ni{p}", f"  {PL[p]} non-interest income", NUM1, f_ni)
    band("Total operating income")
    for p in range(6):
        prow(f"toi{p}", f"  {PL[p]} total op income", NUM1,
             lambda s, c, j, p=p: f"={c}{rmap[f'nii{p}']}+{c}{rmap[f'ni{p}']}")
    band("Operating expenses (cost-to-income, negative)")
    for p in range(6):
        prow(f"opex{p}", f"  {PL[p]} opex", NUM1,
             lambda s, c, j, p=p: f"=-{c}{rmap[f'toi{p}']}*({R[f'cti:{p}']}+({R[f'y5cti:{s}']}-{R['cti:5']}))")
    band("Credit impairment (loss rate x loans, negative)")
    for p in range(6):
        plen = f"{R['stub']}" if p == 0 else "1"
        prow(f"imp{p}", f"  {PL[p]} impairment", NUM1,
             lambda s, c, j, p=p, plen=plen: f"=-{c}{rmap[f'aiea{p}']}*{R['loans_pct']}*({R[f'clr:{p}']}+{R[f'd_clbps:{s}']}/10000)*{plen}")
    band("Cash NPAT")
    for p in range(6):
        prow(f"npat{p}", f"  {PL[p]} NPAT", NUM1,
             lambda s, c, j, p=p: f"=({c}{rmap[f'toi{p}']}+{c}{rmap[f'opex{p}']}+{c}{rmap[f'imp{p}']})*(1-{R['tax']})", bold=(p in (1, 5)))
    for j, sid in enumerate(SIDS):
        for p in range(6):
            R[f"npat:{sid}:{p}"] = f"'P&L Forecast'!{SC[j]}{rmap[f'npat{p}']}"
    ws.column_dimensions["A"].width = 34
    for cc in "CDEFGH": ws.column_dimensions[cc].width = 12

    # ---------------- Valuation (DDM + ROE-fade terminal) ----------------
    ws = wb.create_sheet("Valuation")
    ws.cell(1, 1, "Valuation — dividend discount + ROE-fade terminal, §15.7 equity bridge (AUD m)").font = HDR
    hr = 3; ws.cell(hr, 1, "Line").font = BOLD
    for j, (sid, nm) in enumerate(SCEN):
        ws.cell(hr, 3 + j, nm).font = BOLD
    vr = [hr + 1]; vmap = {}
    def vrow(key, label, fmt, fn, bold=False):
        ws.cell(vr[0], 1, label)
        if bold: ws.cell(vr[0], 1).font = BOLD
        for j, sid in enumerate(SIDS):
            c = ws.cell(vr[0], 3 + j, fn(sid, SC[j], j)); c.number_format = fmt
            if bold: c.font = BOLD
        vmap[key] = vr[0]; vr[0] += 1
    # dividends per period
    for p in range(6):
        vrow(f"div{p}", f"Dividends {PL[p]}", NUM1,
             lambda s, c, j, p=p: f"={R[f'npat:{s}:{p}']}*{R['payout']}")
    # discount factor per period (Ke)
    for p in range(6):
        mt = f"{R['stub']}/2" if p == 0 else f"{R['stub']}+{p}-0.5"
        vrow(f"pv{p}", f"PV dividend {PL[p]}", NUM1,
             lambda s, c, j, p=p, mt=mt: f"={c}{vmap[f'div{p}']}/(1+{R['ke']})^({mt})")
    vrow("pvexp", "PV of explicit dividends", NUM1,
         lambda s, c, j: "=" + "+".join(f"{c}{vmap[f'pv{p}']}" for p in range(6)), bold=True)
    vrow("ret", "Retained earnings (sum NPAT - sum div)", NUM1,
         lambda s, c, j: "=" + "+".join(f"{R[f'npat:{s}:{p}']}" for p in range(6)) + "-(" + "+".join(f"{c}{vmap[f'div{p}']}" for p in range(6)) + ")")
    vrow("ceq", "Closing book equity", NUM0,
         lambda s, c, j: f"={R['book_eq']}+{c}{vmap['ret']}")
    vrow("tv", "Terminal value (equity x (ROE-g)/(Ke-g))", NUM0,
         lambda s, c, j: f"={c}{vmap['ceq']}*({R[f'troe:{s}']}-{R[f'tg:{s}']})/({R['ke']}-{R[f'tg:{s}']})")
    vrow("pvtv", "PV of terminal value", NUM0,
         lambda s, c, j: f"={c}{vmap['tv']}/(1+{R['ke']})^({R['stub']}+{R['horizon']})")
    vrow("claim", "Total value of equity claim", NUM0,
         lambda s, c, j: f"={c}{vmap['pvexp']}+{c}{vmap['pvtv']}", bold=True)
    vrow("ord", "Less AT1/NCI/treasury -> ordinary equity", NUM0,
         lambda s, c, j: f"={c}{vmap['claim']}-{R['at1']}-{R['nci']}-{R['treas']}", bold=True)
    vrow("vps", "Value per share (AUD)", NUM1 + "00",
         lambda s, c, j: f"={c}{vmap['ord']}/{R['shares']}", bold=True)
    for j, sid in enumerate(SIDS):
        R[f"vps:{sid}"] = f"'Valuation'!{SC[j]}{vmap['vps']}"
        R[f"ord:{sid}"] = f"'Valuation'!{SC[j]}{vmap['ord']}"
    ws.column_dimensions["A"].width = 40
    for cc in "CDEFGH": ws.column_dimensions[cc].width = 12

    # ---------------- Comparability & beta (bank peers) ----------------
    ws = wb.create_sheet("Comparability & beta")
    ws.cell(1, 1, "Beta triangulation — bank peers (measured betas; peer-cluster selection §15.2(c))").font = HDR
    ws.cell(2, 1, "ANZ excluded (institutional-dilution outlier); MQG excluded (different archetype). Selected = comparable-cluster midpoint.").font = NOTE
    hr = 4
    for j, t in enumerate(["Peer", "Measured beta", "Role"]):
        ws.cell(hr, 1 + j, t).font = BOLD
    peers = [("CBA", 0.80, "Comparable"), ("NAB", 0.72, "Comparable"), ("WBC", 0.73, "Subject"),  # ssot-allow: peer reference betas
             ("ANZ", 0.57, "Excluded (outlier)"), ("MQG", 0.88, "Excluded (archetype)")]  # ssot-allow: peer reference betas
    rr = hr + 1; incl = []
    for nm, bv, role in peers:
        ws.cell(rr, 1, nm)
        c = ws.cell(rr, 2, bv); c.fill = YELLOW; c.font = BLUEFONT; c.number_format = NUM1 + "0"
        ws.cell(rr, 3, role)
        if role in ("Comparable", "Subject"): incl.append(f"B{rr}")
        rr += 1
    rr += 1
    ws.cell(rr, 1, "Comparable-cluster median").font = BOLD
    ws.cell(rr, 2, "=MEDIAN(" + ",".join(incl) + ")").number_format = NUM1 + "0"; ws.cell(rr, 2).font = BOLD; rr += 1
    ws.cell(rr, 1, "Beta selected (used in Ke, from Assumptions)").font = BOLD
    ws.cell(rr, 2, f"={R['beta']}").number_format = NUM1 + "0"; ws.cell(rr, 2).font = BOLD
    ws.column_dimensions["A"].width = 38
    for cc in "BC": ws.column_dimensions[cc].width = 16

    # ---------------- Multiples (bank: P/E, P/B) ----------------
    ws = wb.create_sheet("Multiples")
    ws.cell(1, 1, "Bank multiples — P/E and P/B on the model's value vs the market").font = HDR
    mp = cfg.get("market", 0)
    rr = 3
    ws.cell(rr, 1, "Market price"); c = ws.cell(rr, 2, mp); c.fill = YELLOW; c.font = BLUEFONT; c.number_format = NUM1 + "0"; mpr = rr; rr += 1
    ws.cell(rr, 1, "FY27 NPAT (Muddle Through)"); ws.cell(rr, 2, f"={R['npat:muddle_through:1']}").number_format = NUM0; np1 = rr; rr += 1
    ws.cell(rr, 1, "EPS FY27 = NPAT / shares"); ws.cell(rr, 2, f"=B{np1}/{R['shares']}").number_format = NUM1 + "00"; eps = rr; rr += 1
    ws.cell(rr, 1, "Book value per share = equity / shares"); ws.cell(rr, 2, f"={R['book_eq']}/{R['shares']}").number_format = NUM1 + "00"; bvps = rr; rr += 1
    rr += 1
    ws.cell(rr, 1, "Market P/E (price / FY27 EPS)"); ws.cell(rr, 2, f"=B{mpr}/B{eps}").number_format = MULT; rr += 1
    ws.cell(rr, 1, "Market P/B (price / book value per share)"); ws.cell(rr, 2, f"=B{mpr}/B{bvps}").number_format = MULT; rr += 1
    ws.cell(rr, 1, "Model P/E (DDM value / FY27 EPS)").font = BOLD
    ws.cell(rr, 2, f"={R['vps:muddle_through']}/B{eps}").number_format = MULT; rr += 1
    ws.cell(rr, 1, "Model P/B (DDM value / book value per share)").font = BOLD
    ws.cell(rr, 2, f"={R['vps:muddle_through']}/B{bvps}").number_format = MULT; rr += 1
    ws.cell(rr, 1, "Cross-check: DDM value per share (Muddle Through)").font = BOLD
    ws.cell(rr, 2, f"={R['vps:muddle_through']}").number_format = NUM1 + "00"; ws.cell(rr, 2).font = BOLD
    ws.column_dimensions["A"].width = 44; ws.column_dimensions["B"].width = 14

    # ---------------- Five forces (company-position offsets) ----------------
    ws = wb.create_sheet("Five forces")
    ws.cell(1, 1, "Porter's five forces \u2192 company-position offsets (feed the AIEA/NIM chain)").font = HDR
    ws.cell(2, 1, "Industry rating is the archetype baseline; the offset is WBC's position. Net offsets link to Assumptions and drive the chain.").font = NOTE
    hr = 4
    for j, t in enumerate(["Force", "Industry rating", "WBC position", "Assessed impact"]):
        ws.cell(hr, 1 + j, t).font = BOLD
    rr = hr + 1
    for row in (cfg.get("_forces", {}) or {}).get("rows", []):
        ws.cell(rr, 1, row[0]); ws.cell(rr, 2, row[1])
        ws.cell(rr, 3, row[3] if len(row) > 3 else "")
        ws.cell(rr, 4, row[4] if len(row) > 4 else "")
        rr += 1
    rr += 1
    ws.cell(rr, 1, "Net NIM offset (deposit-franchise supplier power)").font = BOLD
    ws.cell(rr, 2, f"={R['off_nim']}").number_format = PCT3; ws.cell(rr, 2).font = BOLD; rr += 1
    ws.cell(rr, 1, "Net AIEA-growth offset (rivalry / cost-to-income gap)").font = BOLD
    ws.cell(rr, 2, f"={R['off_aiea']}").number_format = PCT2; ws.cell(rr, 2).font = BOLD; rr += 1
    ws.cell(rr, 1, "These are the company_offset rows in the AIEA/NIM chain \u2014 same values.").font = NOTE
    ws.column_dimensions["A"].width = 40
    for cc in ("B", "C", "D"): ws.column_dimensions[cc].width = 20

    # ---------------- Scenarios summary ----------------
    ws = wb.create_sheet("Scenarios")
    ws.cell(1, 1, "Scenario summary — value per share by world (links to Valuation)").font = HDR
    hr = 3
    for j, t in enumerate(["Scenario", "Value/share", "Ordinary equity", "vs market"]):
        ws.cell(hr, 1 + j, t).font = BOLD
    rr = hr + 1
    for sid, nm in SCEN:
        ws.cell(rr, 1, nm)
        ws.cell(rr, 2, f"={R[f'vps:{sid}']}").number_format = NUM1 + "00"
        ws.cell(rr, 3, f"={R[f'ord:{sid}']}").number_format = NUM0
        ws.cell(rr, 4, f"={R[f'vps:{sid}']}/{mp}-1").number_format = PCT2
        rr += 1
    ws.column_dimensions["A"].width = 26
    for cc in "BCD": ws.column_dimensions[cc].width = 15

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ======================================================================
# CSL segment-FCFF workbook (M3): bottom-up multi-segment FCFF, Ke discount
# ======================================================================

def build_csl_workbook_bytes(cfg=None):
    """Full audited CSL segment-FCFF workbook, formula-linked, tying the engine (MT USD 134.52 / AUD 203.83)."""
    from vcc_valuations.translator import (
        load_inputs as _li, resolve_normalised_baseline as _rnb,
        build_segment_inputs_from_data as _sgi,
    )
    if cfg is None:
        import json as _json
        _cfp = _ROOT / "ui_prototypes" / "_generator" / "cfgs_gen.json"
        cfg = _json.load(open(_cfp))["csl"] if _cfp.exists() else {}
    inp = _li(_ROOT, "muddle_through", "biopharmaceuticals", "csl")
    nb = inp["company_raw"]["normalised_baseline"]; sf = nb["segment_fcff"]
    coe = _rnb(inp)["cost_of_equity_build"]
    fin_segs = inp["financials"]["segments"]
    corp = sf["corporate"]; drv = sf["drivers"]; anch = sf["anchors"]
    SC = sf["by_scenario"]
    SIDS = [sid for sid, _ in SCEN]
    SEGS = [(s["segment"], s["segment"].split("_")[-1], s["fy25_revenue"], s["fy25_or_margin"]) for s in fin_segs]
    YRS = ["FY25", "FY26", "FY27", "FY28", "FY29", "FY30", "FY31"]  # index 0..6

    wb = Workbook(); wb.calculation.fullCalcOnLoad = True
    R = {}

    # ---------------- Assumptions ----------------
    ws = wb.active; ws.title = "Assumptions"; ws.sheet_properties.tabColor = "FFD966"
    r = [1]
    def head(t): ws.cell(r[0], 1, t).font = HDR; r[0] += 1
    def sub(t): ws.cell(r[0], 1, t).font = SUB; r[0] += 1
    def note(t): ws.cell(r[0], 1, t).font = NOTE; r[0] += 1
    def line(label, value, key, fmt=None):
        ws.cell(r[0], 1, label); c = ws.cell(r[0], 2, value); c.fill = YELLOW; c.font = BLUEFONT
        if fmt: c.number_format = fmt
        R[key] = f"'Assumptions'!$B${r[0]}"; r[0] += 1
    head("CSL — Assumptions (yellow cells are the only inputs; segment FCFF / M3)")
    note("Every other sheet links here by formula. Sourced from the production segment engine's data.")
    r[0] += 1
    sub("Cost of equity (single Ke; applied to FCFF, no WACC)")
    line("Risk-free rate Rf", coe["risk_free_rate"], "rf", PCT2)
    line("Equity risk premium ERP", coe["equity_risk_premium"], "erp", PCT2)
    line("Beta (peer-triangulated)", coe["beta"], "beta", NUM1 + "0")
    r[0] += 1
    sub("Shared drivers & anchors")
    line("Capex / revenue", drv["capex_pct_revenue"], "capex", PCT2)
    line("D&A / revenue", drv["da_pct_revenue"], "da", PCT2)
    line("Terminal capex / revenue", drv["terminal_capex_pct_revenue"], "tcapex", PCT2)
    line("Working-capital change / revenue change", drv["wc_change_pct_revenue_change"], "wc", PCT2)
    line("Corporate / unallocated FY25", corp["unallocated_fy25"], "corp0", NUM0)
    line("Corporate growth p.a.", corp["unallocated_growth"], "corpg", PCT2)
    line("Net debt at anchor", anch["net_debt"], "nd", NUM0)
    line("Restructuring cash to come", anch["restructuring_cash_to_come"], "restr", NUM0)
    line("Shares outstanding (m)", anch["shares_outstanding_m"], "shares", NUM1)
    line("FX (AUD per USD)", sf["fx_aud_per_usd"], "fx", NUM1 + "000")
    line("Market price (AUD)", cfg.get("market", 0), "mkt", NUM1 + "0")
    r[0] += 1
    sub("Segment FY25 anchors (USD m; operating-result margin)")
    for sid, stem, rev0, marg0 in SEGS:
        line(f"{stem.title()} FY25 revenue", rev0, f"rev0:{stem}", NUM0)
        line(f"{stem.title()} FY25 OR margin", marg0, f"marg0:{stem}", PCT2)
    r[0] += 1
    # per-scenario driver tables (scenarios as columns C..H)
    def scen_hdr():
        hr = r[0]; ws.cell(hr, 1, "Driver").font = BOLD
        for j, (sid, nm) in enumerate(SCEN):
            cc = ws.cell(hr, 3 + j, nm); cc.font = BOLD; cc.alignment = Alignment(horizontal="right", wrap_text=True)
        r[0] += 1
    def scen_row(label, key, fn, fmt):
        ws.cell(r[0], 1, label)
        for j, sid in enumerate(SIDS):
            c = ws.cell(r[0], 3 + j, fn(sid)); c.fill = YELLOW; c.font = BLUEFONT; c.number_format = fmt
            R[f"{key}:{sid}"] = f"'Assumptions'!${_col(3 + j)}${r[0]}"
        r[0] += 1
    sub("Behring revenue growth path by scenario (FY26..FY31)")
    scen_hdr()
    for y in range(6):
        scen_row(f"  Behring growth {YRS[y+1]}", f"behring_g{y}",
                 lambda sid, y=y: SC[sid]["behring_growth"][y], PCT2)
    sub("Segment CAGR / margin uplift / terminal by scenario")
    scen_hdr()
    scen_row("Seqirus revenue CAGR", "seqirus_cagr", lambda sid: SC[sid]["seqirus_cagr"], PCT2)
    scen_row("Vifor revenue CAGR", "vifor_cagr", lambda sid: SC[sid]["vifor_cagr"], PCT2)
    scen_row("Behring margin uplift (cum FY31)", "up_behring", lambda sid: SC[sid]["behring_margin_uplift"], PCT2)
    scen_row("Seqirus margin uplift (cum FY31)", "up_seqirus", lambda sid: SC[sid]["seqirus_margin_uplift"], PCT2)
    scen_row("Vifor margin uplift (cum FY31)", "up_vifor", lambda sid: SC[sid]["vifor_margin_uplift"], PCT2)
    scen_row("Terminal EBIT margin", "tmargin", lambda sid: SC[sid]["terminal_ebit_margin"], PCT2)
    scen_row("Terminal growth g", "tg", lambda sid: SC[sid]["terminal_growth"], PCT2)
    scen_row("Effective tax rate", "tax", lambda sid: SC[sid]["tax_rate"], PCT2)
    ws.column_dimensions["A"].width = 42
    for cc in "BCDEFGH": ws.column_dimensions[cc].width = 13

    # ---------------- Ke build ----------------
    ws = wb.create_sheet("Ke build")
    ws.cell(1, 1, "Cost of equity (applied to FCFF; no WACC)").font = HDR
    ws.cell(3, 1, "Re = Rf + beta x ERP"); c = ws.cell(3, 2, f"={R['rf']}+{R['beta']}*{R['erp']}"); c.number_format = PCT3; c.font = BOLD
    R["ke"] = "'Ke build'!$B$3"; ws.column_dimensions["A"].width = 26

    # ---------------- Segment FCFF (scenarios as columns) ----------------
    ws = wb.create_sheet("Segment FCFF")
    ws.cell(1, 1, "Bottom-up segment FCFF, six scenarios (USD m). Revenue x margin -> group EBIT -> unlevered FCFF -> per share.").font = HDR
    hr = 3; ws.cell(hr, 1, "Line").font = BOLD
    COLS = [_col(3 + j) for j in range(6)]
    for j, (sid, nm) in enumerate(SCEN):
        ws.cell(hr, 3 + j, nm).font = BOLD
    rr = [hr + 1]; m = {}
    def band(t): ws.cell(rr[0], 1, t).font = SUB; rr[0] += 1
    def row(key, label, fmt, fn, bold=False):
        ws.cell(rr[0], 1, label)
        if bold: ws.cell(rr[0], 1).font = BOLD
        for j, sid in enumerate(SIDS):
            c = ws.cell(rr[0], 3 + j, fn(sid, COLS[j], j)); c.number_format = fmt
            if bold: c.font = BOLD
        m[key] = rr[0]; rr[0] += 1
    # per-segment revenue (FY25..FY31)
    for sid_seg, stem, rev0, marg0 in SEGS:
        band(f"{stem.title()} revenue")
        for p in range(7):
            def f_rev(s, c, j, p=p, stem=stem):
                if p == 0: return f"={R[f'rev0:{stem}']}"
                if stem == "behring":
                    g = R[f'behring_g{p-1}:{s}']
                else:
                    g = R[f'{stem}_cagr:{s}']
                return f"={c}{m[f'{stem}_rev{p-1}']}*(1+{g})"
            row(f"{stem}_rev{p}", f"  {YRS[p]}", NUM0, f_rev)
    # per-segment operating result (rev x margin-glide)
    for sid_seg, stem, rev0, marg0 in SEGS:
        band(f"{stem.title()} operating result")
        for p in range(7):
            def f_or(s, c, j, p=p, stem=stem):
                marg = f"({R[f'marg0:{stem}']}+{R[f'up_{stem}:{s}']}*{p}/6)"
                return f"={c}{m[f'{stem}_rev{p}']}*{marg}"
            row(f"{stem}_or{p}", f"  {YRS[p]}", NUM0, f_or)
    # group rollup
    band("Group")
    for p in range(7):
        row(f"grev{p}", f"  Total revenue {YRS[p]}", NUM0,
            lambda s, c, j, p=p: "=" + "+".join(f"{c}{m[f'{st[1]}_rev{p}']}" for st in SEGS))
    for p in range(7):
        row(f"gor{p}", f"  Segment OR {YRS[p]}", NUM0,
            lambda s, c, j, p=p: "=" + "+".join(f"{c}{m[f'{st[1]}_or{p}']}" for st in SEGS))
    for p in range(7):
        row(f"corp{p}", f"  Corporate {YRS[p]}", NUM0,
            lambda s, c, j, p=p: f"={R['corp0']}*(1+{R['corpg']})^{p}")
    for p in range(7):
        row(f"ebit{p}", f"  Group EBIT {YRS[p]}", NUM0,
            lambda s, c, j, p=p: f"={c}{m[f'gor{p}']}+{c}{m[f'corp{p}']}")
    # FCFF FY26..FY31 (p=1..6)
    band("Free cash flow to firm (FY26..FY31)")
    for p in range(1, 7):
        def f_fcff(s, c, j, p=p):
            nopat = f"{c}{m[f'ebit{p}']}*(1-{R[f'tax:{s}']})"
            da = f"{c}{m[f'grev{p}']}*{R['da']}"
            capex = f"{c}{m[f'grev{p}']}*{R['capex']}"
            wc = f"({c}{m[f'grev{p}']}-{c}{m[f'grev{p-1}']})*{R['wc']}"
            return f"={nopat}+{da}-{capex}-{wc}"
        row(f"fcff{p}", f"  FCFF {YRS[p]}", NUM1, f_fcff)
    # discount + PV (FY27..FY31 explicit)
    for p in range(2, 7):
        mt = p - 1.5
        row(f"pv{p}", f"  PV FCFF {YRS[p]}", NUM1,
            lambda s, c, j, p=p, mt=mt: f"={c}{m[f'fcff{p}']}/(1+{R['ke']})^{mt}")
    row("pvexp", "PV explicit (FY27-FY31)", NUM1,
        lambda s, c, j: "=" + "+".join(f"{c}{m[f'pv{p}']}" for p in range(2, 7)), bold=True)
    # terminal
    row("tfcff", "Terminal FCFF", NUM1,
        lambda s, c, j: f"={c}{m['grev6']}*({R[f'tmargin:{s}']}*(1-{R[f'tax:{s}']})+{R['da']}-{R['tcapex']}-{R[f'tg:{s}']}*{R['wc']})*(1+{R[f'tg:{s}']})")
    row("tv", "Terminal value", NUM0,
        lambda s, c, j: f"={c}{m['tfcff']}/({R['ke']}-{R[f'tg:{s}']})")
    row("pvtv", "PV terminal", NUM0,
        lambda s, c, j: f"={c}{m['tv']}/(1+{R['ke']})^5")
    row("ev", "Enterprise value", NUM0,
        lambda s, c, j: f"={c}{m['pvexp']}+{c}{m['pvtv']}", bold=True)
    row("eq", "Equity value (less net debt, restructuring)", NUM0,
        lambda s, c, j: f"={c}{m['ev']}-{R['nd']}-{R['restr']}", bold=True)
    row("vusd", "Value per share USD", NUM1 + "00",
        lambda s, c, j: f"={c}{m['eq']}/{R['shares']}", bold=True)
    row("vaud", "Value per share AUD", NUM1 + "00",
        lambda s, c, j: f"={c}{m['vusd']}*{R['fx']}", bold=True)
    for j, sid in enumerate(SIDS):
        R[f"vusd:{sid}"] = f"'Segment FCFF'!{COLS[j]}{m['vusd']}"
        R[f"vaud:{sid}"] = f"'Segment FCFF'!{COLS[j]}{m['vaud']}"
        R[f"ev:{sid}"] = f"'Segment FCFF'!{COLS[j]}{m['ev']}"
    ws.column_dimensions["A"].width = 40
    for cc in "CDEFGH": ws.column_dimensions[cc].width = 12

    # ---------------- Comparability & beta ----------------
    import beta_data as _bd
    _cb = _bd.CSL; _idx = _cb["indexDefault"]; _win = _cb["windowDefault"]
    ws = wb.create_sheet("Comparability & beta")
    ws.cell(1, 1, "Beta triangulation — biopharma peers (measured betas; peer-cluster selection)").font = HDR
    ws.cell(2, 1, "Defensive-healthcare cluster. Measured CSL beta rejected (AUD-listed price vs AUD index for a USD earner). Selected = cluster midpoint.").font = NOTE
    hr = 4
    for j, t in enumerate(["Peer", "Measured beta", "Role"]):
        ws.cell(hr, 1 + j, t).font = BOLD
    rr = hr + 1; incl = []
    for c in _cb["comparables"]:
        be = c["data"][_idx][_win]["beta"]
        ws.cell(rr, 1, c["name"])
        cc = ws.cell(rr, 2, be); cc.fill = YELLOW; cc.font = BLUEFONT; cc.number_format = NUM1 + "0"
        ws.cell(rr, 3, "Comparable" if c.get("selected") else "Excluded")
        if c.get("selected"): incl.append(f"B{rr}")
        rr += 1
    rr += 1
    ws.cell(rr, 1, "Comparable-cluster median").font = BOLD
    ws.cell(rr, 2, "=MEDIAN(" + ",".join(incl) + ")").number_format = NUM1 + "0"; ws.cell(rr, 2).font = BOLD; rr += 1
    ws.cell(rr, 1, "Beta selected (used in Ke, from Assumptions)").font = BOLD
    ws.cell(rr, 2, f"={R['beta']}").number_format = NUM1 + "0"; ws.cell(rr, 2).font = BOLD; rr += 1
    ws.cell(rr, 1, _cb["subject"].get("measuredNote", "")[:150]).font = NOTE
    ws.column_dimensions["A"].width = 32
    for c in ("B", "C"): ws.column_dimensions[c].width = 16

    # ---------------- Multiples (EV/EBITDA, EV/EBIT: model vs market, USD) ----------------
    ws = wb.create_sheet("Multiples")
    ws.cell(1, 1, "Multiples \u2014 model EV vs the market-implied EV on FY27 earnings (USD m)").font = HDR
    ws.cell(2, 1, "CSL is USD-functional; the market price (AUD) is converted at the reporting-date FX. Unlevered model, so EV multiples (not P/E).").font = NOTE
    _mtcol = COLS[1]  # Muddle Through column
    rr = 3
    ws.cell(rr, 1, "Group EBIT FY27 (Muddle Through)"); ws.cell(rr, 2, f"='Segment FCFF'!{_mtcol}{m['ebit2']}").number_format = NUM0; _e = rr; rr += 1
    ws.cell(rr, 1, "Group revenue FY27"); ws.cell(rr, 2, f"='Segment FCFF'!{_mtcol}{m['grev2']}").number_format = NUM0; _rev = rr; rr += 1
    ws.cell(rr, 1, "D&A FY27 (revenue x D&A%)"); ws.cell(rr, 2, f"=B{_rev}*{R['da']}").number_format = NUM0; _da = rr; rr += 1
    ws.cell(rr, 1, "EBITDA FY27 (EBIT + D&A)"); ws.cell(rr, 2, f"=B{_e}+B{_da}").number_format = NUM0; _ebda = rr; rr += 1
    rr += 1
    ws.cell(rr, 1, "Model enterprise value (from Segment FCFF)").font = BOLD
    ws.cell(rr, 2, f"={R['ev:muddle_through']}").number_format = NUM0; _mev = rr; rr += 1
    ws.cell(rr, 1, "Model EV / EBITDA"); ws.cell(rr, 2, f"=B{_mev}/B{_ebda}").number_format = '0.0"x"'; rr += 1
    ws.cell(rr, 1, "Model EV / EBIT"); ws.cell(rr, 2, f"=B{_mev}/B{_e}").number_format = '0.0"x"'; rr += 1
    rr += 1
    ws.cell(rr, 1, "Market price (AUD)"); ws.cell(rr, 2, f"={R['mkt']}").number_format = NUM1 + "0"; _mp = rr; rr += 1
    ws.cell(rr, 1, "Market cap (USD) = price/FX x shares"); ws.cell(rr, 2, f"=B{_mp}/{R['fx']}*{R['shares']}").number_format = NUM0; _mc = rr; rr += 1
    ws.cell(rr, 1, "Market EV (USD) = market cap + net debt"); ws.cell(rr, 2, f"=B{_mc}+{R['nd']}").number_format = NUM0; _mkev = rr; rr += 1
    ws.cell(rr, 1, "Market EV / EBITDA"); ws.cell(rr, 2, f"=B{_mkev}/B{_ebda}").number_format = '0.0"x"'; rr += 1
    ws.cell(rr, 1, "Market EV / EBIT"); ws.cell(rr, 2, f"=B{_mkev}/B{_e}").number_format = '0.0"x"'; rr += 1
    rr += 1
    ws.cell(rr, 1, "Cross-check: model value per share (AUD)").font = BOLD
    ws.cell(rr, 2, f"={R['vaud:muddle_through']}").number_format = NUM1 + "00"; ws.cell(rr, 2).font = BOLD
    ws.column_dimensions["A"].width = 42; ws.column_dimensions["B"].width = 14

    # ---------------- Scenarios summary ----------------
    ws = wb.create_sheet("Scenarios")
    ws.cell(1, 1, "Scenario summary — value per share (links to Segment FCFF)").font = HDR
    hr = 3
    for j, t in enumerate(["Scenario", "USD/share", "AUD/share", "EV (USD m)", "vs market"]):
        ws.cell(hr, 1 + j, t).font = BOLD
    rr2 = hr + 1
    for sid, nm in SCEN:
        ws.cell(rr2, 1, nm)
        ws.cell(rr2, 2, f"={R[f'vusd:{sid}']}").number_format = NUM1 + "00"
        ws.cell(rr2, 3, f"={R[f'vaud:{sid}']}").number_format = NUM1 + "00"
        ws.cell(rr2, 4, f"={R[f'ev:{sid}']}").number_format = NUM0
        ws.cell(rr2, 5, f"={R[f'vaud:{sid}']}/{R['mkt']}-1").number_format = PCT2
        rr2 += 1
    ws.column_dimensions["A"].width = 26
    for cc in "BCDE": ws.column_dimensions[cc].width = 13

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


if __name__ == "__main__":
    data = build_dnl_workbook_bytes()
    out = _ROOT / "ui_prototypes" / "_generator" / "_dnl_full.xlsx"
    out.write_bytes(data)
    print("wrote", out, len(data), "bytes")
    wdata = build_wbc_workbook_bytes()
    wout = _ROOT / "ui_prototypes" / "_generator" / "_wbc_full.xlsx"
    wout.write_bytes(wdata)
    print("wrote", wout, len(wdata), "bytes")
    cdata = build_csl_workbook_bytes()
    cout = _ROOT / "ui_prototypes" / "_generator" / "_csl_full.xlsx"
    cout.write_bytes(cdata)
    print("wrote", cout, len(cdata), "bytes")

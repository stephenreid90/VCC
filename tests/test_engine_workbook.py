"""Structural tests for the generated workbooks (batch 6, item 18).

``engine_workbook.py`` is about 1,000 lines that produce the artefact readers
actually open, and it had no direct coverage: the tie tests assert engine-vs-
fixture, not that the builder runs or that what it emits obeys standing rule 1.
A builder can satisfy a value tie while quietly pasting constants where formulas
belong, which is the one thing that rule exists to prevent.

These tests are structural on purpose. They do not check a single valuation
number — that is the tie tests' job — they check that the file is a workbook,
that inputs are inputs, and that everything downstream of them is a formula.
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "ui_prototypes" / "_generator"))

import engine_workbook  # noqa: E402

BUILDERS = {
    "dnl": engine_workbook.build_dnl_workbook_bytes,
    "wbc": engine_workbook.build_wbc_workbook_bytes,
    "csl": engine_workbook.build_csl_workbook_bytes,
}
# Company -> the sheet carrying the valuation build, and the label of the
# working-capital derivation block where one is expected.
VALUATION_SHEET = {"dnl": "DCF build", "wbc": None, "csl": "Segment FCFF"}


@pytest.fixture(scope="module")
def books():
    return {name: load_workbook(io.BytesIO(build())) for name, build in BUILDERS.items()}


@pytest.mark.parametrize("company", list(BUILDERS))
def test_workbook_builds_and_opens(books, company: str) -> None:
    wb = books[company]
    assert "Assumptions" in wb.sheetnames, wb.sheetnames
    assert len(wb.sheetnames) >= 3


@pytest.mark.parametrize("company", list(BUILDERS))
def test_assumptions_sheet_carries_yellow_inputs(books, company: str) -> None:
    """Standing rule 1: inputs are yellow-filled on a dedicated sheet."""
    ws = books[company]["Assumptions"]
    want = engine_workbook.YELLOW.fgColor.rgb
    yellow = [
        c for row in ws.iter_rows() for c in row
        if c.fill is not None and c.fill.patternType == "solid"
        and getattr(c.fill.fgColor, "rgb", None) == want
    ]
    assert len(yellow) > 20, f"{company}: only {len(yellow)} input cells found"
    assert all(not isinstance(c.value, str) or not c.value.startswith("=") for c in yellow), (
        f"{company}: an input cell holds a formula"
    )


@pytest.mark.parametrize("company", [c for c, s in VALUATION_SHEET.items() if s])
def test_valuation_sheet_is_formulas_not_pasted_values(books, company: str) -> None:
    """The rule that matters: no Python-computed constants downstream.

    No NUMBER may appear on the valuation sheet outside column A. Text is fine —
    scenario column headers are text — but a float would mean the builder computed
    something in Python and pasted the answer, which ties perfectly and audits not
    at all.
    """
    ws = books[company][VALUATION_SHEET[company]]
    pasted = [
        f"{c.coordinate}={c.value!r}"
        for row in ws.iter_rows(min_col=2) for c in row
        if isinstance(c.value, (int, float)) and not isinstance(c.value, bool)
    ]
    assert not pasted, f"{company}: pasted values on the valuation sheet: {pasted[:8]}"


@pytest.mark.parametrize("company", ["dnl", "csl"])
def test_working_capital_derivation_is_shown_not_asserted(books, company: str) -> None:
    """The intensity must be derived in the sheet, not typed as one number.

    Guards the shared ``_working_capital_block``: a reader has to be able to see
    the clean-year balance-sheet lines, the average and the rounding, because
    that is the whole argument for the figure.
    """
    ws = books[company]["Assumptions"]
    labels = {
        str(c.value).strip() for row in ws.iter_rows(max_col=1) for c in row
        if c.value is not None
    }
    assert "Working capital — derived intensity (never an input)" in labels
    assert "Average of clean years" in labels
    assert "Applied working-capital intensity" in labels

    applied = next(
        r for r in ws.iter_rows(max_col=2)
        if str(r[0].value).strip() == "Applied working-capital intensity"
    )
    assert isinstance(applied[1].value, str) and applied[1].value.startswith("="), (
        f"{company}: the applied intensity is typed, not derived"
    )


def test_bank_workbook_declares_no_working_capital_block(books) -> None:
    """WBC is exempt by rule (D-11), so the block must be absent, not zeroed."""
    ws = books["wbc"]["Assumptions"]
    labels = {
        str(c.value).strip() for row in ws.iter_rows(max_col=1) for c in row
        if c.value is not None
    }
    assert "Applied working-capital intensity" not in labels

"""Regenerate the generated-workbook oracles — DNL and CSL, six scenarios each.

Why this exists. ``_recalc.py`` pins the engine against
``dnl_muddle_through_valuation_v6_2026-06-25.xlsx``, a hand-built workbook for
one scenario at a superseded beta. That oracle still tests the engine's
arithmetic in ``capitalise_last_fcff`` mode and is kept, but it cannot check the
working-capital rows or the normalised terminal, because the v6 workbook has
neither.

This script recalculates the workbook that ``engine_workbook.build_dnl_workbook_
bytes()`` generates. That workbook is formula-only (standing rule 1): the yellow
Assumptions cells are the sole inputs and every other cell is an Excel formula.
Recalculating it in LibreOffice therefore evaluates the model independently of
the Python engine, and the fixture this writes is a genuine oracle rather than a
restatement of engine output. It regenerates from the data files on every run,
so unlike a hand-built workbook it cannot silently drift from the register.

Regenerate:
    python tests/dcf/golden/_recalc_generated_workbooks.py

Requires ``soffice`` (LibreOffice). ``test_dnl_workbook_tie.py`` and
``test_csl_workbook_tie.py`` read the committed JSON and need no spreadsheet
tooling.
"""

from __future__ import annotations

import json
import subprocess
import sys
import tempfile
from pathlib import Path

REPO = Path(__file__).resolve().parents[3]
HERE = Path(__file__).resolve().parent
DNL_JSON = HERE / "dnl_workbook_all_scenarios.json"
CSL_JSON = HERE / "csl_workbook_all_scenarios.json"

SCENARIOS = [
    "orderly_convergence",
    "muddle_through",
    "ai_productivity_lag",
    "fragmentation",
    "disorderly_climate_crystallisation",
    "stagflation_persists",
]

# Row labels on the "DCF build" sheet -> the engine attribute they must tie to.
# Per-period rows are matched by their "  <period> <label>" prefix.
PERIODS = ["Stub", "Y1", "Y2", "Y3", "Y4", "Y5"]
VECTORS = {
    "revenue": "revenue",
    "EBIT": "ebit",
    "NOPAT": "nopat",
    "D&A": "da",
    "capex": "capex",
    "change in working capital": "delta_wc",
    "FCFF": "fcff",
    "PV of FCFF": "pv_fcff",
}
SCALARS = {
    "PV of explicit FCFF": "pv_explicit",
    "Terminal value = TFCFF/(WACC-g)": "terminal_value",
    "PV of terminal value": "pv_terminal",
    "Enterprise value (EV)": "enterprise_value",
    "Equity value": "equity_value",
    "Value per share": "value_per_share",
}
TERMINAL_FCFF_LABELS = (
    "Terminal FCFF (normalised reinvestment)",
    "Terminal FCFF = Y5 FCFF x (1+g)",
)


# CSL's "Segment FCFF" sheet is laid out one row per line-year rather than one
# row per line with years across, so it needs its own label map.
CSL_YEARS = ["FY26", "FY27", "FY28", "FY29", "FY30", "FY31"]
CSL_VECTORS = {
    "group_revenue": "  Total revenue {y}",
    "group_ebit": "  Group EBIT {y}",
    "wc_change": "  Change in working capital {y}",
    "fcff": "  FCFF {y}",
}
CSL_SCALARS = {
    "PV explicit (FY27-FY31)": "pv_explicit",
    "Terminal FCFF": "terminal_fcff",
    "Terminal value": "terminal_value",
    "PV terminal": "pv_terminal",
    "Enterprise value": "enterprise_value",
    "Equity value (less net debt, restructuring)": "equity_value",
    "Value per share USD": "value_per_share_usd",
    "Value per share AUD": "value_per_share_aud",
}


def _recalc(src: Path, workdir: Path) -> Path:
    prof = workdir / "prof"
    (prof / "user").mkdir(parents=True, exist_ok=True)
    (prof / "user" / "registrymodifications.xcu").write_text(
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<oor:items xmlns:oor="http://openoffice.org/2001/registry" '
        'xmlns:xs="http://www.w3.org/2001/XMLSchema">\n'
        '<item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
        '<prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>\n'
        "</oor:items>\n"
    )
    out = workdir / "out"
    subprocess.run(
        ["soffice", "--headless", f"-env:UserInstallation=file://{prof}",
         "--convert-to", "xlsx:Calc MS Excel 2007 XML", "--outdir", str(out), str(src)],
        check=True, capture_output=True,
    )
    return out / src.name


def _label_rows(ws) -> dict:
    rows: dict[str, int] = {}
    for r in range(1, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if isinstance(label, str):
            rows.setdefault(label.strip(), r)
    return rows


def _extract_dnl(ws) -> dict:
    rows = _label_rows(ws)
    payload = {}
    for j, scenario in enumerate(SCENARIOS):
        col = 3 + j
        vectors = {
            attr: [ws.cell(rows[f"{p} {label}"], col).value for p in PERIODS]
            for label, attr in VECTORS.items()
        }
        scalars = {attr: ws.cell(rows[label], col).value
                   for label, attr in SCALARS.items()}
        for label in TERMINAL_FCFF_LABELS:
            if label in rows:
                scalars["terminal_fcff"] = ws.cell(rows[label], col).value
                break
        payload[scenario] = {"vectors": vectors, "scalars": scalars}
    return payload


def _extract_csl(ws) -> dict:
    rows = _label_rows(ws)
    payload = {}
    for j, scenario in enumerate(SCENARIOS):
        col = 3 + j
        vectors = {
            attr: [ws.cell(rows[tmpl.format(y=y).strip()], col).value for y in CSL_YEARS]
            for attr, tmpl in CSL_VECTORS.items()
        }
        scalars = {attr: ws.cell(rows[label], col).value
                   for label, attr in CSL_SCALARS.items()}
        payload[scenario] = {"vectors": vectors, "scalars": scalars}
    return payload


def main() -> None:
    sys.path.insert(0, str(REPO / "ui_prototypes" / "_generator"))
    import engine_workbook  # noqa: E402
    from openpyxl import load_workbook  # noqa: E402

    jobs = [
        ("dnl", engine_workbook.build_dnl_workbook_bytes, "DCF build", _extract_dnl, DNL_JSON),
        ("csl", engine_workbook.build_csl_workbook_bytes, "Segment FCFF", _extract_csl, CSL_JSON),
    ]
    with tempfile.TemporaryDirectory() as td:
        work = Path(td)
        for name, build, sheet, extract, out in jobs:
            src = work / f"{name}_generated.xlsx"
            src.write_bytes(build())
            ws = load_workbook(_recalc(src, work), data_only=True)[sheet]
            payload = extract(ws)
            out.write_text(json.dumps(payload, indent=2) + "\n")
            print(f"wrote {out.relative_to(REPO)} ({len(payload)} scenarios)")


if __name__ == "__main__":
    main()

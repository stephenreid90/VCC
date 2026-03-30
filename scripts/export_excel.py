"""Export DCF / 3-statement results to Excel workbook."""
import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from models.dcf import DCFModel, DCFAssumptions
from models.three_statement import ThreeStatementModel, ThreeStatementAssumptions


def export_dcf_to_excel(assumptions: DCFAssumptions, path: str = "dcf_output.xlsx"):
    result = DCFModel(assumptions).calculate()
    wb = openpyxl.Workbook()

    # Summary sheet
    ws = wb.active
    ws.title = "DCF Summary"
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(["Metric", "Value (USD)"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    rows = [
        ("Enterprise Value", result.enterprise_value),
        ("Equity Value", result.equity_value),
        ("Price per Share", result.price_per_share),
        ("PV Explicit Period", result.pv_explicit),
        ("PV Terminal Value", result.pv_terminal),
    ]
    for row in rows:
        ws.append(row)

    # FCF projections sheet
    ws2 = wb.create_sheet("FCF Projections")
    ws2.append(["Year", "FCF (USD)"])
    for yr, fcf in enumerate(result.fcf_projections, 1):
        ws2.append([yr, fcf])

    # Sensitivity sheet
    ws3 = wb.create_sheet("Sensitivity")
    ws3.append(["Scenario", "Enterprise Value (USD)"])
    for k, v in result.sensitivity_wacc_growth.items():
        ws3.append([k, v])

    wb.save(path)
    print(f"Saved: {path}")
    return path


if __name__ == "__main__":
    assumptions = DCFAssumptions(
        fcf_base=110e9, fcf_growth_rate=5.0, wacc=7.65,
        terminal_growth=2.5, net_debt=-50e9, shares_outstanding=15_500,
    )
    export_dcf_to_excel(assumptions, "dcf_output.xlsx")

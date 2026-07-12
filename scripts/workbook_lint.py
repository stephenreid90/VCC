#!/usr/bin/env python3
"""workbook_lint.py - integrity checks for VCC company valuation workbooks.

Surfaced from the CSL v1 review (25 June 2026): the workbook stated a 30%
terminal EBIT margin on the Assumptions sheet, but no formula referenced that
cell - the terminal block silently capitalised FY31's ~33% peak margin. With
terminal value ~77% of EV, that one orphaned input was the largest source of
over-valuation in the model.

Two checks implement the proposed methodology v0.7 s11.6 discipline:
  1. ORPHAN-INPUT CHECK (formula-only; always runs). Every Assumptions input
     cell (yellow fill / blue font, literal numeric) must be referenced by at
     least one formula. Structural orphans are errors; documentation-only
     anchors (measured beta, FY25 memos, reconciliation checks, superseded
     inputs) are downgraded to warnings.
  2. TERMINAL-CONTINUITY CHECK (needs values; recalcs via LibreOffice if cached
     values absent). The terminal EBIT margin must actually be wired into the
     terminal block and must not exceed the explicit-period exit-year margin.

Usage: python scripts/workbook_lint.py <workbook.xlsx> [--assumptions-sheet NAME] [--no-recalc]
Exit code is non-zero if any ERROR-level finding is raised (CI-friendly).
"""
from __future__ import annotations
import argparse, os, re, shutil, subprocess, sys, tempfile

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl --break-system-packages")

# Input cells across the VCC workbook set use two palettes: the CSL/WBC
# convention (light-yellow FFF2CC fill / 0066CC blue font) and the older DNL
# convention (FFFF00 fill / 0000FF blue font). Recognise both so the lint does
# not silently detect zero inputs (and false-pass) on a differently-styled book.
YELLOW_SUFFIXES = {"FFF2CC", "FFFF00"}
BLUE_SUFFIXES = {"0066CC", "0000FF"}
STRUCTURAL_HINTS = ["terminal", "discount", "wacc", "cost of equity", "capex",
    "d&a", "tax rate", "net debt", "shares outstanding", "margin uplift",
    "beta selected", "cagr", "risk-free", "equity risk premium", "growth"]
MEMO_HINTS = ["memo", "measured", "sum check", "reconcil", "supersed", "payout", "r&d"]


def _rgb_tail(color):
    try:
        rgb = color.rgb
    except Exception:
        return ""
    return rgb[-6:].upper() if isinstance(rgb, str) else ""


def is_input_cell(cell):
    v = cell.value
    if v is None or isinstance(v, str) or not isinstance(v, (int, float)):
        return False
    fill_yellow = _rgb_tail(cell.fill.fgColor) in YELLOW_SUFFIXES
    font_blue = (cell.font is not None and cell.font.color is not None and
                 _rgb_tail(cell.font.color) in BLUE_SUFFIXES)
    return fill_yellow or font_blue


def classify_orphan(label, note):
    text = ("%s %s" % (label, note)).lower()
    if any(h in text for h in MEMO_HINTS):
        return "WARN"
    if any(h in (label or "").lower() for h in STRUCTURAL_HINTS):
        return "ERROR"
    return "WARN"


def collect_formulas(wb):
    out = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    out.append((ws.title, c.coordinate, c.value))
    return out


def is_referenced(sheet, col, row, formulas):
    q_sheet = re.escape(sheet)
    qualified = re.compile(r"'?%s'?!\s*\$?%s\$?%d(?![0-9])" % (q_sheet, col, row))
    bare = re.compile(r"(?<![A-Za-z0-9_!])\$?%s\$?%d(?![0-9])" % (col, row))
    for f_sheet, _addr, text in formulas:
        if qualified.search(text):
            return True
        if f_sheet == sheet and bare.search(text):
            return True
    return False


def recalc(path):
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        return None
    tmp = tempfile.mkdtemp(prefix="wb_lint_")
    src = os.path.join(tmp, os.path.basename(path))
    shutil.copy(path, src)
    env = dict(os.environ, HOME=tmp)
    try:
        subprocess.run([soffice, "--headless", "--calc", "--convert-to", "xlsx",
                        "--outdir", os.path.join(tmp, "out"), src],
                       check=True, capture_output=True, env=env, timeout=120)
    except Exception:
        return None
    out = os.path.join(tmp, "out", os.path.basename(path))
    return out if os.path.exists(out) else None


def find_labelled_value(ws, needle):
    for r in range(1, ws.max_row + 1):
        lab = ws.cell(r, 1).value
        if isinstance(lab, str) and needle.lower() in lab.lower():
            return r, ws.cell(r, 2).value
    return None, None


def check_lease_basis(wb):
    """AASB 16 lease-basis consistency: if lease liabilities sit on the balance
    sheet, they must be handled in the valuation bridge (included in net debt or
    deducted explicitly per Approach A). Post-AASB 16 EBITDA (rent replaced by
    right-of-use depreciation + lease interest) paired with an ex-lease net debt
    silently overstates equity. Heuristic - WARN, not ERROR."""
    findings = []
    lease_liab = None
    for ws in wb.worksheets:
        for r in range(1, ws.max_row + 1):
            lab = ws.cell(r, 1).value
            if isinstance(lab, str) and re.search(r"lease liabilit", lab, re.I):
                for ci in range(2, min(ws.max_column, 5) + 1):
                    v = ws.cell(r, ci).value
                    if isinstance(v, (int, float)) and v > 0:
                        lease_liab = (ws.title, ws.cell(r, ci).coordinate, v)
                        break
            if lease_liab:
                break
        if lease_liab:
            break
    if not lease_liab:
        findings.append(("INFO", "No lease-liability line found; lease-basis check n/a."))
        return findings
    wired = False
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                lab = c.value
                if isinstance(lab, str) and "lease" in lab.lower() and re.search(
                        r"less:\s*.*lease|add:\s*.*lease|net debt.*includ.*lease|"
                        r"includ.*lease.*net debt|lease.*(=|as)\s*debt|approach a", lab, re.I):
                    wired = True
    if wired:
        findings.append(("INFO", "Lease liabilities (%s!%s = %s) are wired into the bridge / net debt "
                         "(Approach A) - basis consistent." % lease_liab))
    else:
        findings.append(("WARN", "Lease liabilities (%s!%s = %s) are on the balance sheet but not "
                         "explicitly included in net debt or deducted in the equity bridge. If EBITDA/EBIT "
                         "are post-AASB 16, the net-debt side must also carry leases (Approach A) or the "
                         "valuation mixes bases and overstates equity. Confirm treatment." % lease_liab))
    return findings


def lint(path, assumptions_sheet="Assumptions", do_recalc=True):
    findings = []
    wb = openpyxl.load_workbook(path, data_only=False)
    if assumptions_sheet not in wb.sheetnames:
        return [("ERROR", "No '%s' sheet found." % assumptions_sheet)]
    asm = wb[assumptions_sheet]
    formulas = collect_formulas(wb)

    inputs, orphans = 0, []
    terminal_margin_orphaned = False
    for row in asm.iter_rows():
        for c in row:
            if is_input_cell(c):
                inputs += 1
                col = re.match(r"([A-Z]+)", c.coordinate).group(1)
                rn = int(re.match(r"[A-Z]+(\d+)", c.coordinate).group(1))
                if not is_referenced(assumptions_sheet, col, rn, formulas):
                    label = asm.cell(rn, 1).value or "(no label)"
                    note = asm.cell(rn, 3).value or ""
                    level = classify_orphan(label, note)
                    orphans.append((c.coordinate, label, c.value, level))
                    if "terminal ebit margin" in str(label).lower():
                        terminal_margin_orphaned = True
    for addr, label, val, level in orphans:
        tail = "" if level == "ERROR" else " (documentation-only / superseded - review)"
        findings.append((level, "Orphan input %s!%s = %s (%s) not referenced by any formula.%s"
                         % (assumptions_sheet, addr, val, label, tail)))
    n_err = sum(1 for o in orphans if o[3] == "ERROR")
    findings.append(("INFO", "Input cells: %d; referenced: %d; orphan: %d (%d structural)."
                     % (inputs, inputs - len(orphans), len(orphans), n_err)))

    _r, t_margin = find_labelled_value(asm, "terminal ebit margin")
    if (t_margin is None or isinstance(t_margin, str)) and do_recalc:
        rp = recalc(path)
        if rp:
            _r, t_margin = find_labelled_value(
                openpyxl.load_workbook(rp, data_only=True)[assumptions_sheet], "terminal ebit margin")
    if isinstance(t_margin, (int, float)):
        exit_margin = None
        rp = recalc(path) if do_recalc else None
        src_wb = openpyxl.load_workbook(rp, data_only=True) if rp else wb
        for ws in src_wb.worksheets:
            for r in range(1, ws.max_row + 1):
                lab = ws.cell(r, 1).value
                if isinstance(lab, str) and "ebit margin" in lab.lower() and "terminal" not in lab.lower():
                    nums = [ws.cell(r, ci).value for ci in range(2, ws.max_column + 1)
                            if isinstance(ws.cell(r, ci).value, (int, float))]
                    if nums:
                        exit_margin = nums[-1]
        if exit_margin is not None:
            if terminal_margin_orphaned:
                findings.append(("ERROR", "Stated terminal EBIT margin %.1f%% is NOT wired into any "
                    "formula; terminal block instead capitalises exit-year margin %.1f%%. Stated "
                    "assumption is not binding." % (t_margin * 100, exit_margin * 100)))
            elif t_margin > exit_margin + 1e-9:
                findings.append(("ERROR", "Terminal EBIT margin %.1f%% exceeds exit-year margin "
                    "%.1f%% - capitalising a peak into perpetuity." % (t_margin * 100, exit_margin * 100)))
            else:
                findings.append(("INFO", "Terminal margin %.1f%% <= exit-year margin %.1f%% (OK)."
                    % (t_margin * 100, exit_margin * 100)))
        else:
            findings.append(("WARN", "Could not locate an explicit EBIT-margin row for continuity check."))
    else:
        findings.append(("WARN", "Terminal EBIT margin not found / not numeric; continuity check skipped."))

    findings.extend(check_lease_basis(wb))
    return findings


def main():
    ap = argparse.ArgumentParser(description="Lint a VCC valuation workbook for integrity issues.")
    ap.add_argument("workbook")
    ap.add_argument("--assumptions-sheet", default="Assumptions")
    ap.add_argument("--no-recalc", action="store_true")
    args = ap.parse_args()
    findings = lint(args.workbook, args.assumptions_sheet, do_recalc=not args.no_recalc)
    errors = 0
    print("\n=== workbook_lint: %s ===" % os.path.basename(args.workbook))
    for level, msg in findings:
        print("  [%s] %s" % (level, msg))
        if level == "ERROR":
            errors += 1
    print("=== %d error(s) ===\n" % errors)
    sys.exit(1 if errors else 0)


if __name__ == "__main__":
    main()
